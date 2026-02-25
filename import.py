# ============================================================
# IMPORTS — Librerías necesarias para el funcionamiento de la app
# ============================================================
import os           # Manejo de rutas y archivos del sistema operativo
import sqlite3      # Base de datos local SQLite
import calendar     # Utilidades de calendario (calcular días del mes)
import secrets      # Generación segura de tokens y números aleatorios
import hashlib      # Hash criptográfico para contraseñas (PBKDF2)
import smtplib      # Envío de correos mediante protocolo SMTP
import tempfile     # Creación de archivos temporales (para gráficas)
import threading    # Hilos para enviar correos sin bloquear la UI
from email.mime.text import MIMEText                    # Cuerpo del correo en texto/HTML
from email.mime.multipart import MIMEMultipart          # Estructura del correo con múltiples partes
from datetime import datetime, timedelta                # Fechas y diferencia de tiempo

import tkinter as tk                                    # Framework principal de UI de escritorio
from tkinter import ttk, filedialog, messagebox         # Widgets temáticos, diálogos de archivo y alertas

from openpyxl import Workbook                           # Creación de archivos Excel
from openpyxl.styles import Font, PatternFill, Alignment  # Estilos de celdas Excel
from openpyxl.drawing.image import Image as XLImage    # Insertar imágenes en Excel

from reportlab.lib.pagesizes import A4                  # Tamaño de página PDF
from reportlab.platypus import (                        # Componentes de construcción de PDF
    SimpleDocTemplate, Paragraph, Spacer,
    Table as PDFTable, TableStyle, PageBreak, Image as PDFImage
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle  # Estilos de texto PDF
from reportlab.lib.units import cm                      # Unidad de medida en centímetros
from reportlab.lib import colors                        # Colores para el PDF

import matplotlib
matplotlib.use("TkAgg")                                 # Motor de renderizado compatible con Tkinter
from matplotlib.figure import Figure                    # Figura/lienzo de gráficas
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg  # Integra matplotlib en Tkinter
import numpy as np                                      # Arreglos numéricos para las gráficas

# ============================================================
# COMPATIBILIDAD Python 3.14 — Reemplaza ttkbootstrap con ttk nativo
# ============================================================

class Messagebox:
    """
    Clase de compatibilidad que imita la API de ttkbootstrap.Messagebox.
    Permite usar show_info, show_warning, show_error y yesno
    con la misma sintaxis, pero usando los cuadros de diálogo estándar de tkinter.
    """

    @staticmethod
    def show_info(msg, title="Info", parent=None):
        """Muestra un cuadro de diálogo informativo."""
        messagebox.showinfo(title, msg)

    @staticmethod
    def show_warning(msg, title="Aviso", parent=None):
        """Muestra un cuadro de diálogo de advertencia."""
        messagebox.showwarning(title, msg)

    @staticmethod
    def show_error(msg, title="Error", parent=None):
        """Muestra un cuadro de diálogo de error."""
        messagebox.showerror(title, msg)

    @staticmethod
    def yesno(msg, title="Confirmar", parent=None):
        """
        Muestra un cuadro de confirmación Sí/No.
        Devuelve "Yes" o "No" como string para mantener compatibilidad con ttkbootstrap.
        """
        return "Yes" if messagebox.askyesno(title, msg) else "No"


def _patch_widget(cls):
    """
    Parche de compatibilidad para widgets ttk.
    Elimina el parámetro 'bootstyle' (exclusivo de ttkbootstrap)
    antes de llamar al constructor original del widget ttk nativo,
    evitando errores si el código aún usa ese parámetro.
    """
    original = cls.__init__  # Guarda el constructor original del widget

    def new_init(self, *args, **kwargs):
        kwargs.pop("bootstyle", None)  # Elimina 'bootstyle' si viene en los kwargs
        original(self, *args, **kwargs)  # Llama al constructor real sin ese parámetro

    cls.__init__ = new_init  # Reemplaza el constructor por el parchado


# Aplica el parche a todos los widgets ttk que podrían recibir 'bootstyle'
for _cls in (ttk.Button, ttk.Label, ttk.Entry, ttk.Radiobutton,
             ttk.Separator, ttk.Combobox, ttk.Treeview, ttk.Scrollbar):
    _patch_widget(_cls)

# Alias: tb apunta a ttk para mantener compatibilidad con código que usaba "tb.Button", etc.
tb = ttk

# ============================================================
# BASE DE DATOS — Funciones para interactuar con SQLite
# ============================================================

DB_NAME = "presupuesto.db"  # Nombre del archivo de base de datos local


def conectar_db():
    """
    Abre y devuelve una conexión a la base de datos SQLite.
    - timeout=10: espera hasta 10 segundos si la BD está bloqueada.
    - check_same_thread=False: permite usar la conexión desde múltiples hilos.
    - PRAGMA journal_mode=WAL: modo Write-Ahead Logging, evita bloqueos entre lecturas y escrituras.
    """
    conn = sqlite3.connect(DB_NAME, timeout=10, check_same_thread=False)
    conn.execute("PRAGMA journal_mode=WAL")  # Mejora concurrencia en la BD
    return conn


def crear_tabla_movimientos():
    """
    Crea la tabla 'movimientos' si no existe.
    Columnas: id, user_id, tipo (Ingreso/Gasto), concepto, monto, fecha_hora.
    También intenta agregar la columna 'user_id' si la tabla ya existía sin ella
    (compatibilidad hacia atrás con versiones anteriores de la app).
    """
    conn = conectar_db()
    cur = conn.cursor()

    # Crea la tabla con todas las columnas necesarias
    cur.execute("""CREATE TABLE IF NOT EXISTS movimientos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER NOT NULL DEFAULT 0,
        tipo TEXT NOT NULL,
        concepto TEXT NOT NULL,
        monto REAL NOT NULL,
        fecha_hora TEXT NOT NULL)""")
    conn.commit()

    try:
        # Intenta agregar user_id si ya existía la tabla sin esa columna
        cur.execute("ALTER TABLE movimientos ADD COLUMN user_id INTEGER NOT NULL DEFAULT 0")
        conn.commit()
    except Exception:
        pass  # Si ya existe la columna, la excepción se ignora silenciosamente

    conn.close()


def crear_tabla_usuarios():
    """
    Crea la tabla 'usuarios' si no existe.
    Columnas: id, correo (único), nombre, password_hash, salt,
              foto_path (ruta imagen de perfil), reset_code y reset_expira (para recuperación).
    """
    conn = conectar_db()
    cur = conn.cursor()
    cur.execute("""CREATE TABLE IF NOT EXISTS usuarios (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        correo TEXT NOT NULL UNIQUE,
        nombre TEXT NOT NULL,
        password_hash TEXT NOT NULL,
        salt TEXT NOT NULL,
        foto_path TEXT,
        reset_code TEXT,
        reset_expira TEXT)""")
    conn.commit()
    conn.close()


def insertar_movimiento(user_id, tipo, concepto, monto):
    """
    Inserta un nuevo movimiento (ingreso o gasto) en la base de datos.
    La fecha y hora se registran automáticamente en el momento de la inserción.
    """
    conn = conectar_db()
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO movimientos (user_id,tipo,concepto,monto,fecha_hora) VALUES (?,?,?,?,?)",
        (user_id, tipo, concepto, monto, datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    )
    conn.commit()
    conn.close()


def eliminar_movimiento_por_id(mov_id, user_id):
    """
    Elimina un movimiento específico por su ID.
    El filtro por user_id garantiza que un usuario solo pueda borrar sus propios registros.
    """
    conn = conectar_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM movimientos WHERE id=? AND user_id=?", (mov_id, user_id))
    conn.commit()
    conn.close()


def obtener_movimientos_rango(user_id, fecha_inicio, fecha_fin):
    """
    Devuelve todos los movimientos del usuario dentro de un rango de fechas.
    Los resultados vienen ordenados del más reciente al más antiguo (ORDER BY id DESC).
    Retorna lista de tuplas: (id, tipo, concepto, monto, fecha_hora).
    """
    conn = conectar_db()
    cur = conn.cursor()
    cur.execute("""SELECT id,tipo,concepto,monto,fecha_hora FROM movimientos
                   WHERE user_id=? AND fecha_hora>=? AND fecha_hora<=? ORDER BY id DESC""",
                (user_id, fecha_inicio, fecha_fin))
    filas = cur.fetchall()
    conn.close()
    return filas


def calcular_totales_rango(user_id, fecha_inicio, fecha_fin):
    """
    Calcula y devuelve los totales del período: ingresos, gastos y saldo.
    Usa COALESCE para devolver 0 si no hay registros (evita None).
    Retorna una tupla: (total_ingresos, total_gastos, saldo).
    """
    conn = conectar_db()
    cur = conn.cursor()

    # Suma todos los ingresos del período
    cur.execute(
        "SELECT COALESCE(SUM(monto),0) FROM movimientos WHERE user_id=? AND tipo='Ingreso' AND fecha_hora>=? AND fecha_hora<=?",
        (user_id, fecha_inicio, fecha_fin)
    )
    ing = float(cur.fetchone()[0])

    # Suma todos los gastos del período
    cur.execute(
        "SELECT COALESCE(SUM(monto),0) FROM movimientos WHERE user_id=? AND tipo='Gasto' AND fecha_hora>=? AND fecha_hora<=?",
        (user_id, fecha_inicio, fecha_fin)
    )
    gas = float(cur.fetchone()[0])

    conn.close()
    return ing, gas, ing - gas  # Devuelve también el saldo calculado


def meses_disponibles(user_id):
    """
    Devuelve la lista de meses (formato 'YYYY-MM') en los que el usuario
    tiene registros, ordenados del más reciente al más antiguo.
    Se usa para poblar el ComboBox de filtro por mes.
    """
    conn = conectar_db()
    cur = conn.cursor()
    # substr(fecha_hora,1,7) extrae solo los primeros 7 caracteres → "YYYY-MM"
    cur.execute(
        "SELECT DISTINCT substr(fecha_hora,1,7) FROM movimientos WHERE user_id=? ORDER BY 1 DESC",
        (user_id,)
    )
    rows = cur.fetchall()
    conn.close()
    return [r[0] for r in rows if r[0]]  # Filtra posibles valores vacíos


def rango_del_mes(yyyy_mm):
    """
    Dado un mes en formato 'YYYY-MM', devuelve el rango completo de fechas:
    desde el primer día a las 00:00:00 hasta el último día a las 23:59:59.
    Usa calendar.monthrange para obtener el último día del mes correctamente.
    """
    year, month = int(yyyy_mm[:4]), int(yyyy_mm[5:7])
    ultimo = calendar.monthrange(year, month)[1]  # Día final del mes (28/29/30/31)
    return f"{yyyy_mm}-01 00:00:00", f"{yyyy_mm}-{ultimo:02d} 23:59:59"


def eliminar_cuenta(user_id):
    """
    Elimina permanentemente la cuenta del usuario de la tabla 'usuarios'.
    Los movimientos asociados no se eliminan automáticamente (podrían quedar huérfanos).
    """
    conn = conectar_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM usuarios WHERE id=?", (user_id,))
    conn.commit()
    conn.close()


def actualizar_usuario(user_id, nombre, correo, foto_path=None):
    """
    Actualiza los datos de perfil del usuario.
    Si se proporciona foto_path, también actualiza la foto de perfil.
    Si no se pasa foto_path, solo actualiza nombre y correo.
    """
    conn = conectar_db()
    cur = conn.cursor()
    if foto_path is not None:
        cur.execute(
            "UPDATE usuarios SET nombre=?,correo=?,foto_path=? WHERE id=?",
            (nombre, correo, foto_path, user_id)
        )
    else:
        cur.execute(
            "UPDATE usuarios SET nombre=?,correo=? WHERE id=?",
            (nombre, correo, user_id)
        )
    conn.commit()
    conn.close()


def cambiar_password_usuario(user_id, nueva_password):
    """
    Cambia la contraseña del usuario de forma segura.
    Genera un nuevo salt aleatorio y recalcula el hash antes de guardar.
    """
    salt_hex = secrets.token_hex(16)                    # Nuevo salt de 16 bytes en hexadecimal
    pass_hash = _hash_password(nueva_password, salt_hex)  # Hash de la nueva contraseña
    conn = conectar_db()
    cur = conn.cursor()
    cur.execute(
        "UPDATE usuarios SET password_hash=?,salt=? WHERE id=?",
        (pass_hash, salt_hex, user_id)
    )
    conn.commit()
    conn.close()


# ============================================================
# SEGURIDAD — Hashing de contraseñas y autenticación
# ============================================================

def _hash_password(password, salt_hex):
    """
    Genera un hash seguro de la contraseña usando PBKDF2-HMAC-SHA256.
    - salt_hex: cadena hexadecimal que se convierte a bytes.
    - 200_000 iteraciones: hace el hash costoso para resistir ataques de fuerza bruta.
    Devuelve el hash resultante como cadena hexadecimal.
    """
    salt = bytes.fromhex(salt_hex)  # Convierte el salt de hex a bytes
    return hashlib.pbkdf2_hmac("sha256", password.encode(), salt, 200_000).hex()


def crear_usuario(correo, nombre, password, foto_path):
    """
    Registra un nuevo usuario en la base de datos.
    - Normaliza correo a minúsculas y elimina espacios.
    - Valida que los campos no estén vacíos.
    - Genera un salt único y hashea la contraseña antes de guardar.
    Lanza ValueError si faltan datos, o sqlite3.IntegrityError si el correo ya existe.
    """
    correo = correo.strip().lower()  # Normaliza el correo
    nombre = nombre.strip()
    if not correo or not nombre or not password:
        raise ValueError("Faltan datos.")

    salt_hex = secrets.token_hex(16)  # Salt único para este usuario
    conn = conectar_db()
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO usuarios (correo,nombre,password_hash,salt,foto_path) VALUES (?,?,?,?,?)",
        (correo, nombre, _hash_password(password, salt_hex), salt_hex, foto_path)
    )
    conn.commit()
    conn.close()


def verificar_login(correo, password):
    """
    Verifica las credenciales de inicio de sesión.
    - Busca al usuario por correo (normalizado a minúsculas).
    - Hashea la contraseña ingresada con el salt almacenado y la compara.
    Devuelve un diccionario con los datos del usuario si es correcto, o None si falla.
    """
    correo = correo.strip().lower()
    conn = conectar_db()
    cur = conn.cursor()
    cur.execute(
        "SELECT id,correo,nombre,password_hash,salt,foto_path FROM usuarios WHERE correo=?",
        (correo,)
    )
    row = cur.fetchone()
    conn.close()

    if not row:
        return None  # Usuario no encontrado

    user_id, correo, nombre, ph, salt, foto_path = row

    if _hash_password(password, salt) != ph:
        return None  # Contraseña incorrecta

    # Devuelve los datos del usuario autenticado
    return {"id": user_id, "correo": correo, "nombre": nombre, "foto_path": foto_path}


def crear_codigo_reset(correo):
    """
    Genera un código de 6 dígitos para recuperación de contraseña.
    - El código expira en 10 minutos.
    - Se guarda en la tabla usuarios junto con la fecha de expiración.
    - La búsqueda del correo es insensible a mayúsculas/minúsculas.
    Devuelve el código generado, o None si el correo no está registrado.
    """
    correo = correo.strip().lower()
    codigo = f"{secrets.randbelow(1_000_000):06d}"  # Número aleatorio de 6 dígitos con ceros al inicio
    expira = (datetime.now() + timedelta(minutes=10)).strftime("%Y-%m-%d %H:%M:%S")  # Expiración en 10 min

    conn = conectar_db()
    cur = conn.cursor()

    # Busca el correo ignorando diferencias de mayúsculas/minúsculas
    cur.execute("SELECT id, correo FROM usuarios WHERE LOWER(correo)=?", (correo,))
    row = cur.fetchone()
    if not row:
        conn.close()
        return None  # Correo no registrado

    correo_real = row[1]  # Usa el correo exactamente como está guardado en la BD
    cur.execute(
        "UPDATE usuarios SET reset_code=?,reset_expira=? WHERE correo=?",
        (codigo, expira, correo_real)
    )
    conn.commit()
    conn.close()
    return codigo


def resetear_password(correo, codigo, nueva_password):
    """
    Cambia la contraseña del usuario si el código de verificación es válido.
    - Verifica que el código no haya expirado.
    - Compara el código ingresado con el guardado en BD.
    - Si todo es correcto, genera nuevo salt y hash y actualiza la contraseña.
    - Limpia el código y fecha de expiración después de usarlos.
    Devuelve True si fue exitoso, False si el código es inválido o expiró.
    """
    correo = correo.strip().lower()
    conn = conectar_db()
    cur = conn.cursor()
    cur.execute("SELECT reset_code,reset_expira FROM usuarios WHERE correo=?", (correo,))
    row = cur.fetchone()

    if not row or not row[0]:
        conn.close()
        return False  # No existe código de reset para este correo

    reset_code, reset_expira = row

    try:
        # Verifica si el código ya expiró
        if datetime.now() > datetime.strptime(reset_expira, "%Y-%m-%d %H:%M:%S"):
            conn.close()
            return False
    except Exception:
        conn.close()
        return False

    if codigo.strip() != reset_code:
        conn.close()
        return False  # Código incorrecto

    # Genera nueva contraseña hasheada y la guarda, limpiando el código de reset
    salt_hex = secrets.token_hex(16)
    cur.execute(
        "UPDATE usuarios SET password_hash=?,salt=?,reset_code=NULL,reset_expira=NULL WHERE correo=?",
        (_hash_password(nueva_password, salt_hex), salt_hex, correo)
    )
    conn.commit()
    conn.close()
    return True


# ============================================================
# CORREO — Envío de emails HTML mediante Gmail
# ============================================================

# Credenciales del remitente (cuenta de Gmail dedicada para la app)
GMAIL_REMITENTE = "recuperaciocorreo967@gmail.com"
GMAIL_APP_PASS  = "xido uqzf moqk lgxa"  # Contraseña de aplicación de Gmail (no la contraseña real)


def _enviar_html(destinatario, asunto, html):
    """
    Función interna que envía un correo HTML usando SMTP SSL con Gmail.
    - Construye el mensaje MIME con formato HTML y UTF-8.
    - Usa el puerto 465 con SSL para conexión segura.
    Devuelve True si el envío fue exitoso, False si ocurrió algún error.
    """
    try:
        msg = MIMEMultipart("alternative")  # Contenedor de correo multipart
        msg["From"] = GMAIL_REMITENTE
        msg["To"] = destinatario
        msg["Subject"] = asunto
        msg.attach(MIMEText(html, "html", "utf-8"))  # Adjunta el cuerpo HTML

        with smtplib.SMTP_SSL("smtp.gmail.com", 465, timeout=10) as s:
            s.login(GMAIL_REMITENTE, GMAIL_APP_PASS)  # Autentica con Gmail
            s.sendmail(GMAIL_REMITENTE, destinatario, msg.as_string())  # Envía

        return True
    except Exception as e:
        print(f"[Email error] {e}")  # Registra el error en consola sin romper la app
        return False


def enviar_codigo_por_correo(destinatario, codigo):
    """
    Envía un correo HTML con el código de recuperación de contraseña.
    El diseño incluye gradiente púrpura, código en grande y aviso de expiración.
    """
    html = f"""<html><body style="font-family:'Segoe UI',Arial;background:#F3F4F6;padding:40px 0;">
    <table width="100%"><tr><td align="center"><table width="480" style="background:#fff;border-radius:16px;overflow:hidden;box-shadow:0 4px 24px rgba(0,0,0,.1);">
    <tr><td style="background:linear-gradient(135deg,#7C3AED,#4C1D95);padding:36px 40px;text-align:center;">
    <h1 style="color:#fff;font-size:22px;margin:10px 0 4px;">Recuperar contraseña</h1></td></tr>
    <tr><td style="padding:36px 40px;">
    <p style="color:#374151;">Tu código de verificación:</p>
    <div style="background:#F5F3FF;border:2px dashed #7C3AED;border-radius:12px;padding:20px;text-align:center;">
    <span style="font-size:36px;font-weight:bold;letter-spacing:10px;color:#7C3AED;">{codigo}</span></div>
    <p style="color:#92400E;background:#FEF3C7;border-left:4px solid #F59E0B;padding:12px;border-radius:6px;margin-top:16px;">
    ⏱️ Expira en <strong>10 minutos</strong>.</p></td></tr>
    <tr><td style="background:#F9FAFB;padding:16px 40px;text-align:center;">
    <p style="color:#9CA3AF;font-size:12px;">© 2026 Presupuesto Personal</p></td></tr>
    </table></td></tr></table></body></html>"""
    return _enviar_html(destinatario, "💜 Tu código de recuperación — Presupuesto Personal", html)


def enviar_bienvenida(destinatario, nombre):
    """
    Envía un correo de bienvenida al usuario recién registrado.
    Incluye su nombre y correo como confirmación del registro exitoso.
    """
    html = f"""<html><body style="font-family:'Segoe UI',Arial;background:#F3F4F6;padding:40px 0;">
    <table width="100%"><tr><td align="center"><table width="480" style="background:#fff;border-radius:16px;overflow:hidden;box-shadow:0 4px 24px rgba(0,0,0,.1);">
    <tr><td style="background:linear-gradient(135deg,#7C3AED,#4C1D95);padding:36px 40px;text-align:center;">
    <h1 style="color:#fff;font-size:24px;">¡Bienvenido a Presupuesto Personal!</h1></td></tr>
    <tr><td style="padding:36px 40px;">
    <p style="color:#374151;font-size:16px;">Hola <strong>{nombre}</strong> 👋</p>
    <p style="color:#374151;">Tu cuenta está lista. Ya puedes registrar tus ingresos y gastos.</p>
    <p style="color:#9CA3AF;font-size:13px;">Correo: <strong>{destinatario}</strong></p></td></tr>
    <tr><td style="background:#F9FAFB;padding:16px 40px;text-align:center;">
    <p style="color:#9CA3AF;font-size:12px;">© 2026 Presupuesto Personal</p></td></tr>
    </table></td></tr></table></body></html>"""
    return _enviar_html(destinatario, "💜 ¡Bienvenido a Presupuesto Personal!", html)


def enviar_confirmacion_cambio(destinatario, nombre):
    """
    Envía un correo de seguridad notificando que la contraseña fue cambiada.
    Incluye fecha y hora del cambio, y una advertencia por si no fue el usuario.
    """
    ahora = datetime.now().strftime("%d/%m/%Y %H:%M")  # Fecha y hora actuales formateadas
    html = f"""<html><body style="font-family:'Segoe UI',Arial;background:#F3F4F6;padding:40px 0;">
    <table width="100%"><tr><td align="center"><table width="480" style="background:#fff;border-radius:16px;overflow:hidden;box-shadow:0 4px 24px rgba(0,0,0,.1);">
    <tr><td style="background:linear-gradient(135deg,#059669,#065F46);padding:36px 40px;text-align:center;">
    <h1 style="color:#fff;font-size:22px;">¡Contraseña actualizada!</h1></td></tr>
    <tr><td style="padding:36px 40px;">
    <p style="color:#374151;">Hola <strong>{nombre}</strong>, tu contraseña fue actualizada el <strong>{ahora}</strong>.</p>
    <p style="color:#065F46;background:#ECFDF5;border-left:4px solid #10B981;padding:12px;border-radius:6px;">
    🔒 Si no fuiste tú, cambia tu contraseña de inmediato.</p></td></tr>
    <tr><td style="background:#F9FAFB;padding:16px 40px;text-align:center;">
    <p style="color:#9CA3AF;font-size:12px;">© 2026 Presupuesto Personal</p></td></tr>
    </table></td></tr></table></body></html>"""
    return _enviar_html(destinatario, "✅ Contraseña actualizada — Presupuesto Personal", html)


# ============================================================
# LOGIN — Ventana de inicio de sesión
# ============================================================

class LoginWindow(tk.Toplevel):
    """
    Ventana modal de inicio de sesión.
    Permite ingresar credenciales, abrir el registro o recuperar contraseña.
    Al autenticarse correctamente, llama al callback on_success con los datos del usuario.
    """

    def __init__(self, master, on_success):
        super().__init__(master)
        self.on_success = on_success  # Función a llamar cuando el login sea exitoso

        self.title("Presupuesto — Iniciar sesión")
        self.geometry("480x340")
        self.resizable(False, False)

        # Variables de formulario enlazadas a los campos de entrada
        self.correo_var = tk.StringVar()
        self.pass_var   = tk.StringVar()

        # Contenedor principal con padding interno
        wrap = ttk.Frame(self, padding=28)
        wrap.pack(fill="both", expand=True)

        # Títulos de bienvenida
        ttk.Label(wrap, text="💜 Bienvenido",
                  font=("Segoe UI", 20, "bold")).pack(pady=(0, 4))
        ttk.Label(wrap, text="Ingresa tus credenciales para continuar",
                  font=("Segoe UI", 10)).pack(pady=(0, 16))

        # Formulario de correo y contraseña en grid
        form = ttk.Frame(wrap)
        form.pack(fill="x")

        ttk.Label(form, text="📧 Correo", font=("Segoe UI", 10)).grid(row=0, column=0, sticky="w", pady=4)
        correo_e = ttk.Entry(form, textvariable=self.correo_var, width=38, font=("Segoe UI", 10))
        correo_e.grid(row=0, column=1, padx=8, pady=4)

        ttk.Label(form, text="🔒 Contraseña", font=("Segoe UI", 10)).grid(row=1, column=0, sticky="w", pady=4)
        pass_e = ttk.Entry(form, textvariable=self.pass_var, width=38, show="•", font=("Segoe UI", 10))
        pass_e.grid(row=1, column=1, padx=8, pady=4)
        pass_e.bind("<Return>", lambda e: self._login())  # Enter dispara el login

        # Botones de acción
        btns = ttk.Frame(wrap)
        btns.pack(fill="x", pady=(16, 6))

        ttk.Button(btns, text="Entrar",             width=12, command=self._login).pack(side="left", padx=4)
        ttk.Button(btns, text="Crear cuenta",       width=14, command=self._open_register).pack(side="left", padx=4)
        ttk.Button(btns, text="Olvidé mi contraseña",         command=self._open_reset).pack(side="right", padx=4)

        # Centra la ventana en la pantalla
        self.update_idletasks()
        x = (self.winfo_screenwidth()  // 2) - (self.winfo_width()  // 2)
        y = (self.winfo_screenheight() // 2) - (self.winfo_height() // 2)
        self.geometry(f"+{x}+{y}")

        self.grab_set()     # Bloquea interacción con otras ventanas (modal)
        self.focus_force()  # Fuerza el foco a esta ventana
        correo_e.focus()    # Posiciona el cursor en el campo de correo

    def _login(self):
        """
        Valida que los campos no estén vacíos y llama a verificar_login.
        Si las credenciales son correctas, cierra la ventana y llama a on_success.
        Si son incorrectas, muestra error y limpia el campo de contraseña.
        """
        if not self.correo_var.get().strip() or not self.pass_var.get():
            Messagebox.show_warning("Completa todos los campos.", "Atención", parent=self)
            return

        user = verificar_login(self.correo_var.get(), self.pass_var.get())

        if not user:
            Messagebox.show_error("Correo o contraseña incorrectos.", "Error", parent=self)
            self.pass_var.set("")  # Limpia la contraseña por seguridad
            return

        self.grab_release()  # Libera el bloqueo modal
        self.destroy()       # Cierra la ventana de login
        self.on_success(user)  # Notifica al padre con los datos del usuario

    def _open_register(self):
        """Abre la ventana de registro. Libera el grab para que la nueva ventana pueda tomarlo."""
        self.grab_release()
        RegisterWindow(self, on_close=lambda: self.grab_set())  # Al cerrar el registro, retoma el grab

    def _open_reset(self):
        """Abre la ventana de recuperación de contraseña."""
        ResetWindow(self)


# ============================================================
# REGISTRO — Ventana para crear nueva cuenta
# ============================================================

class RegisterWindow(tk.Toplevel):
    """
    Ventana modal para registrar un nuevo usuario.
    Valida los campos, crea el usuario en la BD y envía correo de bienvenida.
    """

    def __init__(self, master, on_close=None):
        super().__init__(master)
        self._on_close = on_close  # Callback que se ejecuta al cerrar esta ventana

        self.title("Crear cuenta")
        self.geometry("520x420")
        self.resizable(False, False)

        # Variables del formulario
        self.correo_var = tk.StringVar()
        self.nombre_var = tk.StringVar()
        self.pass1_var  = tk.StringVar()
        self.pass2_var  = tk.StringVar()
        self.foto_path  = None  # Ruta opcional de imagen de perfil

        wrap = ttk.Frame(self, padding=24)
        wrap.pack(fill="both", expand=True)

        ttk.Label(wrap, text="Crear cuenta", font=("Segoe UI", 16, "bold")).pack(anchor="w", pady=(0, 12))

        # Formulario con los 4 campos del registro
        form = ttk.Labelframe(wrap, text="Datos de registro", padding=14)
        form.pack(fill="x")

        for i, (lbl, var, oculto) in enumerate([
            ("Nombre",            self.nombre_var, False),
            ("Correo",            self.correo_var, False),
            ("Contraseña",        self.pass1_var,  True),
            ("Repite contraseña", self.pass2_var,  True)
        ]):
            ttk.Label(form, text=lbl).grid(row=i, column=0, sticky="w", padx=6, pady=5)
            kw = {"show": "•"} if oculto else {}  # Oculta texto en campos de contraseña
            ttk.Entry(form, textvariable=var, width=42, **kw).grid(row=i, column=1, padx=6, pady=5)

        # Sección para elegir foto de perfil
        foto_row = ttk.Frame(wrap)
        foto_row.pack(fill="x", pady=10)
        self.lbl_foto = ttk.Label(foto_row, text="Foto de perfil: (opcional)")
        self.lbl_foto.pack(side="left")
        ttk.Button(foto_row, text="Elegir imagen", command=self._pick_foto).pack(side="right")

        # Etiqueta de estado para mostrar mensajes de error o éxito
        self.lbl_estado = tk.Label(wrap, text="", font=("Segoe UI", 10, "bold"), fg="#059669")
        self.lbl_estado.pack(anchor="w", pady=(0, 4))

        ttk.Button(wrap, text="Crear usuario", width=18, command=self._crear).pack(pady=4)

        # Centra la ventana en pantalla
        self.update_idletasks()
        x = (self.winfo_screenwidth()  // 2) - (self.winfo_width()  // 2)
        y = (self.winfo_screenheight() // 2) - (self.winfo_height() // 2)
        self.geometry(f"+{x}+{y}")

        self.transient(master)  # Asocia esta ventana como hija de master
        self.grab_set()
        self.focus_force()
        self.protocol("WM_DELETE_WINDOW", self._cerrar)  # Intercepta el cierre con la X

    def _cerrar(self):
        """Libera el grab, ejecuta el callback de cierre y destruye la ventana."""
        self.grab_release()
        if self._on_close:
            self._on_close()
        self.destroy()

    def _pick_foto(self):
        """Abre un diálogo para seleccionar una imagen de perfil y guarda la ruta."""
        path = filedialog.askopenfilename(
            title="Elegir foto",
            filetypes=[("Imágenes", "*.png *.jpg *.jpeg"), ("Todos", "*.*")]
        )
        if path:
            self.foto_path = path
            self.lbl_foto.config(text=f"Foto: {os.path.basename(path)}")

    def _crear(self):
        """
        Valida los datos del formulario y crea la cuenta de usuario.
        - Verifica que nombre y correo no estén vacíos.
        - Verifica que las contraseñas coincidan y tengan mínimo 6 caracteres.
        - Llama a crear_usuario() y envía correo de bienvenida en un hilo separado.
        """
        nombre = self.nombre_var.get().strip()
        correo = self.correo_var.get().strip()
        p1 = self.pass1_var.get()
        p2 = self.pass2_var.get()

        if not nombre or not correo:
            self.lbl_estado.config(text="⚠️ Nombre y correo son obligatorios.", fg="#DC2626")
            return
        if p1 != p2:
            self.lbl_estado.config(text="❌ Las contraseñas no coinciden.", fg="#DC2626")
            return
        if len(p1) < 6:
            self.lbl_estado.config(text="⚠️ Contraseña muy corta (mínimo 6).", fg="#DC2626")
            return

        self.lbl_estado.config(text="⏳ Creando cuenta…", fg="#D97706")
        self.update()  # Fuerza la actualización visual de la etiqueta

        try:
            crear_usuario(correo, nombre, p1, self.foto_path)
        except sqlite3.IntegrityError:
            self.lbl_estado.config(text="❌ Ese correo ya está registrado.", fg="#DC2626")
            return
        except Exception as e:
            self.lbl_estado.config(text=f"❌ Error: {e}", fg="#DC2626")
            return

        self.lbl_estado.config(text="✅ ¡Cuenta creada!", fg="#059669")
        self.update()

        # Envía el correo de bienvenida en un hilo separado para no bloquear la UI
        threading.Thread(target=enviar_bienvenida, args=(correo, nombre), daemon=True).start()
        self.after(1200, self._cerrar)  # Cierra la ventana después de 1.2 segundos


# ============================================================
# RECUPERAR CONTRASEÑA — Flujo de 3 pasos
# ============================================================

class ResetWindow(tk.Toplevel):
    """
    Ventana de recuperación de contraseña en 3 fases:
    1. Ingresar correo → enviar código por email.
    2. Ingresar el código recibido → verificar.
    3. Crear nueva contraseña → guardar.
    """

    def __init__(self, master):
        super().__init__(master)
        self.title("Recuperar contraseña")
        self.resizable(False, False)
        self._correo = ""  # Correo del usuario que quiere recuperar su contraseña

        # Variables de los 4 campos del flujo de reset
        self.correo_var = tk.StringVar()
        self.codigo_var = tk.StringVar()
        self.pass1_var  = tk.StringVar()
        self.pass2_var  = tk.StringVar()

        # Contenedor reutilizable que se limpia en cada fase
        self._wrap = ttk.Frame(self, padding=28)
        self._wrap.pack(fill="both", expand=True)

        self._mostrar_fase1()  # Comienza en la fase 1
        self.grab_set()

    def _limpiar(self):
        """Destruye todos los widgets del contenedor para preparar la siguiente fase."""
        for w in self._wrap.winfo_children():
            w.destroy()

    def _recentrar(self, geo):
        """Ajusta el tamaño de la ventana y la recentra en pantalla."""
        self.geometry(geo)
        self.update_idletasks()
        x = (self.winfo_screenwidth()  // 2) - (self.winfo_width()  // 2)
        y = (self.winfo_screenheight() // 2) - (self.winfo_height() // 2)
        self.geometry(f"{geo}+{x}+{y}")

    def _mostrar_fase1(self):
        """Fase 1: Muestra el campo para ingresar el correo y el botón de enviar código."""
        self._limpiar()
        self._recentrar("580x280")

        ttk.Label(self._wrap, text="🔐  Recuperar contraseña",
                  font=("Segoe UI", 16, "bold")).pack(anchor="w", pady=(0, 14))

        f = ttk.Labelframe(self._wrap, text="Paso 1 — Ingresa tu correo registrado", padding=16)
        f.pack(fill="x")

        row = ttk.Frame(f)
        row.pack(fill="x")
        ttk.Label(row, text="Correo:", font=("Segoe UI", 10)).pack(side="left", padx=(0, 8))
        correo_e = ttk.Entry(row, textvariable=self.correo_var, width=36, font=("Segoe UI", 10))
        correo_e.pack(side="left", expand=True, fill="x")
        correo_e.focus()

        # Etiqueta de estado para esta fase
        self.lbl_f1 = tk.Label(self._wrap, text="", font=("Segoe UI", 10, "bold"), fg="#D97706")
        self.lbl_f1.pack(anchor="w", pady=(10, 4))

        btns = ttk.Frame(self._wrap)
        btns.pack(fill="x", pady=(6, 0))

        self.btn_enviar = ttk.Button(btns, text="📨  Enviar código", width=22, command=self._enviar_codigo)
        self.btn_enviar.pack(side="left", padx=(0, 8))
        ttk.Button(btns, text="Cancelar", command=self.destroy).pack(side="left")

    def _enviar_codigo(self):
        """
        Genera el código de reset y lo envía al correo.
        Deshabilita el botón durante el envío para evitar doble clic.
        Si el envío falla, muestra el código en pantalla como fallback.
        """
        correo = self.correo_var.get().strip()
        if not correo:
            self.lbl_f1.config(text="⚠️ Escribe tu correo.", fg="#DC2626")
            return

        codigo = crear_codigo_reset(correo)
        if not codigo:
            self.lbl_f1.config(text="❌ Correo no registrado.", fg="#DC2626")
            return

        self.btn_enviar.config(state="disabled")  # Desactiva el botón durante el envío
        self.lbl_f1.config(text="⏳ Enviando…", fg="#D97706")
        self.update()

        ok = enviar_codigo_por_correo(correo, codigo)  # Envío sincrónico (puede tardar)
        self.btn_enviar.config(state="normal")
        self._correo = correo

        if ok:
            self.lbl_f1.config(text=f"✅ Código enviado a {correo}", fg="#059669")
        else:
            # Fallback: muestra el código en pantalla si el correo falló
            self.lbl_f1.config(text=f"⚠️ No se pudo enviar. Código: {codigo}", fg="#DC2626")

        self.after(900, self._mostrar_fase2)  # Avanza a la fase 2 después de 0.9 segundos

    def _mostrar_fase2(self):
        """Fase 2: Campo para ingresar el código de 6 dígitos recibido por correo."""
        self._limpiar()
        self._recentrar("580x300")

        ttk.Label(self._wrap, text="🔑  Verificar código",
                  font=("Segoe UI", 16, "bold")).pack(anchor="w", pady=(0, 14))

        f = ttk.Labelframe(self._wrap, text="Paso 2 — Ingresa el código recibido", padding=16)
        f.pack(fill="x")

        row = ttk.Frame(f)
        row.pack(fill="x")
        ttk.Label(row, text="Código de 6 dígitos:", font=("Segoe UI", 10)).pack(side="left", padx=(0, 10))
        cod_e = ttk.Entry(row, textvariable=self.codigo_var, width=14,
                          font=("Segoe UI", 16, "bold"), justify="center")
        cod_e.pack(side="left")
        cod_e.focus()
        cod_e.bind("<Return>", lambda e: self._validar_codigo())  # Enter valida el código

        self.lbl_f2 = tk.Label(self._wrap, text="Revisa tu correo e ingresa el código.",
                                font=("Segoe UI", 10), fg="#6B7280")
        self.lbl_f2.pack(anchor="w", pady=(10, 4))

        btns = ttk.Frame(self._wrap)
        btns.pack(fill="x", pady=(10, 0))
        ttk.Button(btns, text="✅  Verificar", width=16, command=self._validar_codigo).pack(side="left", padx=(0, 8))
        ttk.Button(btns, text="← Volver",     width=10, command=self._mostrar_fase1).pack(side="left")

    def _validar_codigo(self):
        """
        Verifica que el código ingresado sea correcto y no haya expirado.
        Si es válido, avanza a la fase 3.
        """
        codigo = self.codigo_var.get().strip()
        if not codigo:
            self.lbl_f2.config(text="⚠️ Ingresa el código.", fg="#DC2626")
            return

        conn = conectar_db()
        cur = conn.cursor()
        cur.execute("SELECT reset_code,reset_expira FROM usuarios WHERE correo=?", (self._correo,))
        row = cur.fetchone()
        conn.close()

        if not row or not row[0]:
            self.lbl_f2.config(text="❌ Código no encontrado.", fg="#DC2626")
            return

        reset_code, reset_expira = row

        try:
            if datetime.now() > datetime.strptime(reset_expira, "%Y-%m-%d %H:%M:%S"):
                self.lbl_f2.config(text="⏰ Código expirado.", fg="#DC2626")
                return
        except Exception:
            self.lbl_f2.config(text="❌ Error al verificar.", fg="#DC2626")
            return

        if codigo != reset_code:
            self.lbl_f2.config(text="❌ Código incorrecto.", fg="#DC2626")
            return

        self.lbl_f2.config(text="✅ Código verificado.", fg="#059669")
        self.after(700, self._mostrar_fase3)  # Avanza a fase 3 tras 0.7 segundos

    def _mostrar_fase3(self):
        """Fase 3: Campos para ingresar y confirmar la nueva contraseña."""
        self._limpiar()
        self._recentrar("580x320")

        ttk.Label(self._wrap, text="🔒  Nueva contraseña",
                  font=("Segoe UI", 16, "bold")).pack(anchor="w", pady=(0, 14))

        f = ttk.Labelframe(self._wrap, text="Paso 3 — Crea tu nueva contraseña", padding=16)
        f.pack(fill="x")

        for i, (lbl, var) in enumerate([("Nueva contraseña:", self.pass1_var), ("Repite contraseña:", self.pass2_var)]):
            row = ttk.Frame(f)
            row.pack(fill="x", pady=6)
            ttk.Label(row, text=lbl, font=("Segoe UI", 10), width=18, anchor="w").pack(side="left")
            e = ttk.Entry(row, textvariable=var, show="•", font=("Segoe UI", 10), width=30)
            e.pack(side="left", expand=True, fill="x")
            if i == 0:
                e.focus()  # Foco en el primer campo
            e.bind("<Return>", lambda ev: self._guardar_password())  # Enter guarda

        self.lbl_f3 = tk.Label(self._wrap, text="Mínimo 6 caracteres.",
                                font=("Segoe UI", 10), fg="#6B7280")
        self.lbl_f3.pack(anchor="w", pady=(10, 4))

        btns = ttk.Frame(self._wrap)
        btns.pack(fill="x", pady=(10, 0))
        ttk.Button(btns, text="💾  Guardar contraseña", width=24, command=self._guardar_password).pack(side="left", padx=(0, 8))
        ttk.Button(btns, text="Cancelar", width=10, command=self.destroy).pack(side="left")

    def _guardar_password(self):
        """
        Valida las contraseñas y llama a resetear_password().
        Envía correo de confirmación de cambio en hilo separado.
        Cierra la ventana al completar exitosamente.
        """
        p1 = self.pass1_var.get()
        p2 = self.pass2_var.get()

        if not p1:
            self.lbl_f3.config(text="⚠️ Escribe una contraseña.", fg="#DC2626")
            return
        if p1 != p2:
            self.lbl_f3.config(text="❌ No coinciden.", fg="#DC2626")
            return
        if len(p1) < 6:
            self.lbl_f3.config(text="⚠️ Mínimo 6 caracteres.", fg="#DC2626")
            return

        ok = resetear_password(self._correo, self.codigo_var.get().strip(), p1)
        if not ok:
            self.lbl_f3.config(text="❌ Código expirado. Empieza de nuevo.", fg="#DC2626")
            return

        # Obtiene el nombre del usuario para personalizar el correo de confirmación
        conn = conectar_db()
        cur = conn.cursor()
        cur.execute("SELECT nombre FROM usuarios WHERE correo=?", (self._correo,))
        row = cur.fetchone()
        conn.close()
        nombre = row[0] if row else self._correo

        self.lbl_f3.config(text="✅ Contraseña actualizada.", fg="#059669")
        self.update()

        # Envía notificación de cambio en segundo plano
        threading.Thread(target=enviar_confirmacion_cambio, args=(self._correo, nombre), daemon=True).start()

        Messagebox.show_info("🎉 ¡Contraseña actualizada!\nYa puedes iniciar sesión.", "¡Listo!", parent=self)
        self.grab_release()
        self.destroy()


# ============================================================
# EDITAR PERFIL — Ventana para modificar datos de la cuenta
# ============================================================

class EditarPerfilWindow(tk.Toplevel):
    """
    Ventana modal para que el usuario edite su nombre, correo, contraseña y foto de perfil.
    También permite eliminar la cuenta permanentemente.
    """

    def __init__(self, master, usuario, on_update):
        super().__init__(master)
        self.usuario   = usuario    # Diccionario con los datos actuales del usuario
        self.on_update = on_update  # Callback llamado al guardar o eliminar la cuenta

        self.title("Editar perfil")
        self.geometry("520x440")
        self.resizable(False, False)

        # Variables inicializadas con los datos actuales del usuario
        self.nombre_var = tk.StringVar(value=usuario["nombre"])
        self.correo_var = tk.StringVar(value=usuario["correo"])
        self.pass1_var  = tk.StringVar()
        self.pass2_var  = tk.StringVar()
        self.foto_path  = usuario.get("foto_path")

        wrap = ttk.Frame(self, padding=24)
        wrap.pack(fill="both", expand=True)

        ttk.Label(wrap, text="✏️  Editar perfil",
                  font=("Segoe UI", 16, "bold")).pack(anchor="w", pady=(0, 14))

        # Formulario con los 4 campos editables
        form = ttk.Labelframe(wrap, text="Información de la cuenta", padding=14)
        form.pack(fill="x")

        for i, (lbl, var, oculto) in enumerate([
            ("Nombre",                    self.nombre_var, False),
            ("Correo",                    self.correo_var, False),
            ("Nueva contraseña (opcional)", self.pass1_var,  True),
            ("Repite contraseña",         self.pass2_var,  True)
        ]):
            ttk.Label(form, text=lbl).grid(row=i, column=0, sticky="w", padx=6, pady=6)
            kw = {"show": "•"} if oculto else {}
            ttk.Entry(form, textvariable=var, width=38, **kw).grid(row=i, column=1, padx=6, pady=6)

        # Sección para cambiar la foto de perfil
        foto_row = ttk.Frame(wrap)
        foto_row.pack(fill="x", pady=10)
        nombre_foto = os.path.basename(self.foto_path) if self.foto_path else "ninguna"
        self.lbl_foto = ttk.Label(foto_row, text=f"Foto actual: {nombre_foto}")
        self.lbl_foto.pack(side="left")
        ttk.Button(foto_row, text="Cambiar foto", command=self._pick_foto).pack(side="right")

        # Botones de acción
        btns = ttk.Frame(wrap)
        btns.pack(fill="x", pady=10)
        ttk.Button(btns, text="Guardar cambios",     command=self._guardar).pack(side="left",  padx=6)
        ttk.Button(btns, text="Cancelar",            command=self.destroy).pack(side="left",  padx=6)
        ttk.Button(btns, text="🗑️ Eliminar cuenta", command=self._eliminar_cuenta).pack(side="right", padx=6)

        self.grab_set()

    def _pick_foto(self):
        """Abre un diálogo para seleccionar una nueva foto de perfil."""
        path = filedialog.askopenfilename(
            title="Elegir foto",
            filetypes=[("Imágenes", "*.png *.jpg *.jpeg"), ("Todos", "*.*")]
        )
        if path:
            self.foto_path = path
            self.lbl_foto.config(text=f"Foto: {os.path.basename(path)}")

    def _eliminar_cuenta(self):
        """
        Elimina permanentemente la cuenta del usuario tras doble confirmación.
        Llama a on_update(None) para que la app principal sepa que debe regresar al login.
        """
        ok1 = Messagebox.yesno("⚠️ ¿Eliminar tu cuenta?\n\nEsta acción no se puede deshacer.", "Eliminar cuenta", parent=self)
        if ok1 != "Yes":
            return

        # Segunda confirmación con nombre y correo visible
        ok2 = Messagebox.yesno(
            f"🚨 CONFIRMACIÓN FINAL\n\n{self.usuario['nombre']} ({self.usuario['correo']})\n\n¿Confirmas?",
            "¿Estás seguro?", parent=self
        )
        if ok2 != "Yes":
            return

        eliminar_cuenta(self.usuario["id"])
        Messagebox.show_info("Cuenta eliminada.", "Listo")
        self.destroy()
        self.on_update(None)  # None indica que la cuenta fue eliminada

    def _guardar(self):
        """
        Guarda los cambios del perfil: nombre, correo, foto y opcionalmente la contraseña.
        Actualiza el diccionario del usuario en memoria y llama a on_update.
        """
        nombre = self.nombre_var.get().strip()
        correo = self.correo_var.get().strip().lower()
        p1     = self.pass1_var.get()
        p2     = self.pass2_var.get()

        if not nombre or not correo:
            Messagebox.show_warning("Nombre y correo son obligatorios.", "Atención", parent=self)
            return

        # Cambia la contraseña solo si se ingresó alguna
        if p1 or p2:
            if p1 != p2:
                Messagebox.show_error("Las contraseñas no coinciden.", "Error", parent=self)
                return
            if len(p1) < 6:
                Messagebox.show_warning("Contraseña muy corta.", "Atención", parent=self)
                return
            cambiar_password_usuario(self.usuario["id"], p1)

        actualizar_usuario(self.usuario["id"], nombre, correo, self.foto_path)
        self.usuario.update({"nombre": nombre, "correo": correo, "foto_path": self.foto_path})

        Messagebox.show_info("¡Perfil actualizado!", "Listo")
        self.on_update(self.usuario)  # Notifica con los datos actualizados
        self.destroy()


# ============================================================
# COLORES — Constantes de color usadas en la interfaz
# ============================================================
GREEN     = "#10B981"  # Verde para ingresos
RED       = "#EF4444"  # Rojo para gastos
BLUE      = "#3B82F6"  # Azul para saldo positivo
BG_HEADER = "#0F0E2A"  # Color de fondo del encabezado (azul muy oscuro)


# ============================================================
# APP PRINCIPAL — Ventana principal de la aplicación
# ============================================================

class App(ttk.Frame):
    """
    Frame principal de la aplicación de presupuesto.
    Contiene el encabezado con perfil, el formulario de registro de movimientos,
    la barra de filtros, la tabla de movimientos y las gráficas.
    """

    def __init__(self, root, usuario):
        super().__init__(root)
        self.root    = root      # Referencia a la ventana raíz de Tkinter
        self.usuario = usuario   # Diccionario con datos del usuario autenticado

        self.root.title("💜 Presupuesto Personal")
        self.root.geometry("1220x800")
        self.root.minsize(1100, 720)  # Tamaño mínimo para que la UI no se rompa

        # Fuentes reutilizables
        self.FONT      = ("Segoe UI", 11)
        self.FONT_BOLD = ("Segoe UI", 11, "bold")

        # Variables de formulario enlazadas a los widgets
        self.tipo_var     = tk.StringVar(value="Ingreso")
        self.concepto_var = tk.StringVar()
        self.monto_var    = tk.StringVar()
        self.mes_var      = tk.StringVar()

        # Estado de animación de las barras
        self._anim_target  = [0, 0, 0]    # Valores finales de la animación
        self._anim_current = [0.0, 0.0, 0.0]  # Valores actuales durante la animación
        self._anim_step    = 0             # Paso actual de la animación
        self._anim_id      = None          # ID del after() activo para cancelarlo si es necesario

        # Estado de animación del hover en el avatar
        self._hover_anim_id    = None
        self._pill_bg          = BG_HEADER  # Color de fondo actual del contenedor del avatar
        self._menu_abierto     = None       # Referencia al menú de perfil si está abierto
        self._menu_dot_abierto = None       # Referencia al menú ⋮ si está abierto

        self.pack(fill="both", expand=True)  # El frame principal ocupa toda la ventana
        self._build_ui()   # Construye todos los widgets de la interfaz
        self._refresh()    # Carga los datos iniciales
        self._auto_refresh()  # Inicia el ciclo de auto-actualización cada 60 segundos

    def _build_ui(self):
        """Construye la estructura general de la interfaz: encabezado, formulario, tabla y gráficas."""

        # ── Encabezado ──────────────────────────────────────
        header = tk.Frame(self, bg=BG_HEADER, height=62)
        header.pack(fill="x", side="top")
        header.pack_propagate(False)  # Mantiene la altura fija del encabezado

        # Canvas para el avatar/pill del usuario (foto circular + nombre)
        self._pill_canvas = tk.Canvas(header, width=170, height=46,
                                      bg=BG_HEADER, highlightthickness=0, cursor="hand2")
        self._pill_canvas.pack(side="left", padx=10, pady=8)
        self._foto_img_ref = None  # Referencia de imagen para evitar que el GC la elimine
        self._menu_abierto = None
        self._dibujar_pill_canvas()  # Dibuja el avatar inicial

        # Eventos del avatar: clic abre menú, hover anima el fondo
        for ev, fn in [("<Button-1>", lambda e: self._toggle_perfil_menu()),
                       ("<Enter>",    lambda e: self._hover_perfil(True)),
                       ("<Leave>",    lambda e: self._hover_perfil(False))]:
            self._pill_canvas.bind(ev, fn)

        # Título central de la app
        tk.Label(header, text="💜  Presupuesto Personal",
                 bg=BG_HEADER, fg="#C4B5FD",
                 font=("Segoe UI", 14, "bold")).pack(side="left", expand=True)

        # Botón de menú ⋮ (tres puntos) en el extremo derecho
        self._menu_dot_abierto = None
        menu_btn = tk.Label(header, text="⋮", bg=BG_HEADER, fg="#7C6FCD",
                            font=("Segoe UI", 22, "bold"), cursor="hand2", padx=16)
        menu_btn.pack(side="right")
        menu_btn.bind("<Button-1>", self._toggle_dot_menu)
        menu_btn.bind("<Enter>", lambda e: menu_btn.config(fg="#C4B5FD"))  # Resalta al pasar el mouse
        menu_btn.bind("<Leave>", lambda e: menu_btn.config(fg="#7C6FCD"))  # Restaura al salir

        # ── Cuerpo principal ──────────────────────────────────
        body = ttk.Frame(self, padding=(14, 10))
        body.pack(fill="both", expand=True)

        self._build_form(body)         # Formulario de nuevo movimiento
        ttk.Separator(body, orient="horizontal").pack(fill="x", pady=6)  # Línea separadora
        self._build_filter_bar(body)   # Barra de filtros y tarjetas de resumen

        # Panel dividido: tabla a la izquierda, gráficas a la derecha
        panels = ttk.Frame(body)
        panels.pack(fill="both", expand=True)
        panels.columnconfigure(0, weight=55)  # Tabla ocupa 55% del ancho
        panels.columnconfigure(1, weight=45)  # Gráficas ocupan 45%
        panels.rowconfigure(0, weight=1)

        self._build_table(panels)   # Tabla de movimientos
        self._build_charts(panels)  # Área de gráficas

    def _build_form(self, parent):
        """
        Construye el formulario para registrar un nuevo movimiento.
        Campos: tipo (Ingreso/Gasto), concepto y monto. Botón Registrar.
        """
        form = ttk.Labelframe(parent, text="➕  Nuevo movimiento", padding=10)
        form.pack(fill="x")

        # Selector de tipo mediante radio buttons
        ttk.Label(form, text="Tipo:", font=self.FONT).grid(row=0, column=0, sticky="w", padx=6)
        ttk.Radiobutton(form, text="Ingreso", variable=self.tipo_var, value="Ingreso").grid(row=0, column=1, padx=6)
        ttk.Radiobutton(form, text="Gasto",   variable=self.tipo_var, value="Gasto").grid(row=0, column=2, padx=6)

        # Campo de concepto
        ttk.Label(form, text="Concepto:", font=self.FONT).grid(row=0, column=3, sticky="w", padx=(20, 6))
        ttk.Entry(form, textvariable=self.concepto_var, width=28, font=self.FONT).grid(row=0, column=4, padx=6)

        # Campo de monto
        ttk.Label(form, text="Monto $:", font=self.FONT).grid(row=0, column=5, sticky="w", padx=(12, 6))
        ttk.Entry(form, textvariable=self.monto_var, width=14, font=self.FONT).grid(row=0, column=6, padx=6)

        # Botón de registro
        ttk.Button(form, text="Registrar", command=self._registrar).grid(row=0, column=7, padx=16)

    def _build_filter_bar(self, parent):
        """
        Construye la barra de filtros y las tarjetas de resumen (Ingresos, Gastos, Saldo).
        El ComboBox permite filtrar por mes; el botón Ver todo muestra todos los registros.
        """
        filt = ttk.Frame(parent)
        filt.pack(fill="x", pady=(0, 6))

        # ComboBox de selección de mes
        ttk.Label(filt, text="Filtrar por mes:", font=self.FONT).pack(side="left", padx=(0, 8))
        self.cb_mes = ttk.Combobox(filt, textvariable=self.mes_var, state="readonly", width=14)
        self.cb_mes.pack(side="left")
        self.cb_mes.bind("<<ComboboxSelected>>", lambda e: self._refresh())  # Refresca al cambiar mes

        ttk.Button(filt, text="Ver todo", command=self._ver_todo).pack(side="left", padx=8)

        # Tarjetas de resumen al lado derecho de la barra
        card_frame = ttk.Frame(filt)
        card_frame.pack(side="right")

        self.lbl_ing   = tk.Label(card_frame, text="Ingresos: $0",  fg=GREEN, bg="#F0FDF4", font=("Segoe UI", 11, "bold"), padx=10, pady=4)
        self.lbl_gas   = tk.Label(card_frame, text="Gastos: $0",    fg=RED,   bg="#FEF2F2", font=("Segoe UI", 11, "bold"), padx=10, pady=4)
        self.lbl_saldo = tk.Label(card_frame, text="Saldo: $0",     fg=BLUE,  bg="#EFF6FF", font=("Segoe UI", 11, "bold"), padx=10, pady=4)

        for lbl in (self.lbl_ing, self.lbl_gas, self.lbl_saldo):
            lbl.pack(side="left", padx=8)

    def _build_table(self, parent):
        """
        Construye el Treeview (tabla) de movimientos con columnas:
        ID, Tipo, Concepto, Monto, Fecha/Hora.
        Ingresos se muestran en verde y gastos en rojo.
        """
        frame = ttk.Labelframe(parent, text="📋  Movimientos", padding=6)
        frame.grid(row=0, column=0, sticky="nsew", padx=(0, 6))

        cols = ("ID", "Tipo", "Concepto", "Monto", "Fecha / Hora")
        self.tree = ttk.Treeview(frame, columns=cols, show="headings", selectmode="browse")

        for col, w in zip(cols, [50, 80, 220, 110, 160]):
            self.tree.heading(col, text=col)
            self.tree.column(col, width=w, anchor="center" if col != "Concepto" else "w")

        # Tags para colorear filas según tipo
        self.tree.tag_configure("Ingreso", foreground=GREEN)
        self.tree.tag_configure("Gasto",   foreground=RED)

        # Scrollbar vertical sincronizada con el Treeview
        scroll = ttk.Scrollbar(frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scroll.set)
        scroll.pack(side="right", fill="y")
        self.tree.pack(fill="both", expand=True)

    def _build_charts(self, parent):
        """
        Construye el área de gráficas con matplotlib embebido en Tkinter.
        Contiene dos subplots: barras (resumen del período) y líneas (historial mensual).
        """
        frame = ttk.Labelframe(parent, text="📊  Gráficas", padding=6)
        frame.grid(row=0, column=1, sticky="nsew")

        BG = "#0F172A"  # Color de fondo oscuro para las gráficas

        # Figura matplotlib con dos subplots apilados verticalmente
        self.fig = Figure(figsize=(5.6, 5.4), dpi=96, facecolor=BG)
        self.fig.subplots_adjust(hspace=0.46, top=0.94, bottom=0.09, left=0.13, right=0.97)
        self.ax_bar  = self.fig.add_subplot(211)  # Gráfica superior: barras
        self.ax_line = self.fig.add_subplot(212)  # Gráfica inferior: líneas históricas

        for ax in (self.ax_bar, self.ax_line):
            ax.set_facecolor(BG)  # Fondo oscuro para cada subplot

        # Canvas que integra la figura de matplotlib en el widget de Tkinter
        self.canvas_fig = FigureCanvasTkAgg(self.fig, master=frame)
        self.canvas_fig.get_tk_widget().pack(fill="both", expand=True)

    # ── Avatar / Pill del usuario ─────────────────────────

    def _dibujar_pill_canvas(self):
        """
        Dibuja el avatar del usuario en el Canvas del encabezado.
        Si PIL está disponible y hay foto válida, muestra la imagen circular.
        Si no, muestra un círculo con la inicial del nombre.
        También muestra el nombre abreviado al lado del avatar.
        """
        c = self._pill_canvas
        c.delete("all")          # Limpia el canvas antes de redibujar
        c.config(bg=self._pill_bg)

        AV = 40   # Tamaño del avatar en píxeles
        ax0 = 6   # Posición X del avatar
        ay0 = 3   # Posición Y del avatar

        fp = self.usuario.get("foto_path")  # Ruta de la foto de perfil

        try:
            from PIL import Image, ImageTk, ImageDraw
            if not fp or not os.path.exists(fp):
                raise FileNotFoundError

            # Carga, redimensiona y recorta la foto en forma circular
            img = Image.open(fp).convert("RGBA").resize((AV, AV), Image.LANCZOS)
            mask = Image.new("L", (AV, AV), 0)
            ImageDraw.Draw(mask).ellipse((0, 0, AV, AV), fill=255)
            img.putalpha(mask)  # Aplica la máscara circular con transparencia

            self._foto_img_ref = ImageTk.PhotoImage(img)  # Guarda referencia para evitar GC
            c.create_image(ax0 + AV // 2, ay0 + AV // 2, image=self._foto_img_ref)
            c.create_oval(ax0, ay0, ax0 + AV, ay0 + AV, outline="#7C3AED", width=2)  # Borde púrpura

        except Exception:
            # Fallback: círculo con la inicial del nombre del usuario
            self._foto_img_ref = None
            c.create_oval(ax0, ay0, ax0 + AV, ay0 + AV, fill="#2A1660", outline="#7C3AED", width=2)
            c.create_text(ax0 + AV // 2, ay0 + AV // 2,
                          text=self.usuario.get("nombre", "?")[0].upper(),
                          fill="white", font=("Segoe UI", 14, "bold"))

        # Muestra el nombre truncado a 14 caracteres al lado del avatar
        nombre_corto = self.usuario.get("nombre", "Usuario")
        nombre_corto = nombre_corto[:14] + "…" if len(nombre_corto) > 14 else nombre_corto
        c.create_text(ax0 + AV + 8, ay0 + AV // 2,
                      text=nombre_corto, fill="#EDE9FE",
                      font=("Segoe UI", 9, "bold"), anchor="w")

    def _hover_perfil(self, entering):
        """
        Anima suavemente el color de fondo del avatar entre el color base y uno más claro.
        Se llama al entrar (entering=True) y salir (entering=False) del canvas.
        Usa easing cuadrático (ease-in-out) para suavizar la transición en 12 pasos.
        """
        if self._hover_anim_id:
            try:
                self.root.after_cancel(self._hover_anim_id)  # Cancela animación previa
            except Exception:
                pass
            self._hover_anim_id = None

        C0 = (15, 14, 42)   # Color base del header en RGB
        C1 = (28, 25, 65)   # Color de hover (más claro) en RGB
        STEPS = 12           # Número de pasos de la animación

        try:
            h = self._pill_bg.lstrip("#")
            cur = (int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16))  # Color actual en RGB
        except Exception:
            cur = C0

        target = C1 if entering else C0  # Color destino según si entra o sale

        def _step(i):
            """Calcula e interpola el color en el paso i con easing cuadrático."""
            t = i / STEPS
            t = t * (2 - t)  # Función de easing ease-in-out
            r = int(cur[0] + (target[0] - cur[0]) * t)
            g = int(cur[1] + (target[1] - cur[1]) * t)
            b = int(cur[2] + (target[2] - cur[2]) * t)
            hx = "#{:02x}{:02x}{:02x}".format(r, g, b)
            self._pill_bg = hx
            self._pill_canvas.config(bg=hx)
            if i < STEPS:
                self._hover_anim_id = self.root.after(12, lambda: _step(i + 1))  # Siguiente paso en 12ms
            else:
                self._hover_anim_id = None

        _step(1)  # Inicia la animación en el paso 1

    def _toggle_perfil_menu(self):
        """Alterna la visibilidad del menú de perfil. Si está abierto lo cierra, si no lo abre."""
        if self._menu_abierto:
            try:
                if self._menu_abierto.winfo_exists():
                    self._on_menu_close(self._menu_abierto)
                    return
            except Exception:
                pass
        self._show_perfil_menu()

    def _show_perfil_menu(self):
        """
        Crea y muestra el menú desplegable del perfil del usuario.
        Contiene: foto de perfil grande, nombre, correo, opción de editar perfil y cerrar sesión.
        Se posiciona justo debajo del avatar en el encabezado.
        Se cierra al perder el foco, presionar Escape o al mover/redimensionar la ventana.
        """
        BG  = "#0D0C22"   # Fondo oscuro del menú
        BG2 = "#13112E"   # Fondo ligeramente más claro para la sección superior
        SEP = "#1E1C42"   # Color de las líneas separadoras
        ACC = "#7C3AED"   # Color acento púrpura

        menu = tk.Toplevel(self.root)
        menu.overrideredirect(True)         # Sin borde ni barra de título del SO
        menu.attributes("-topmost", True)   # Siempre encima de otras ventanas

        # Posiciona el menú justo debajo del canvas del avatar
        px = self._pill_canvas.winfo_rootx()
        py = self._pill_canvas.winfo_rooty() + self._pill_canvas.winfo_height() + 6
        menu.geometry(f"+{px}+{py}")

        self._menu_abierto = menu  # Guarda referencia para poder cerrarlo

        # Contenedor con borde acento
        outer = tk.Frame(menu, bg=BG, highlightbackground=ACC, highlightthickness=1)
        outer.pack(fill="both", expand=True)

        # Sección superior: avatar grande + nombre + correo + link cambiar foto
        top = tk.Frame(outer, bg=BG2, padx=16, pady=14)
        top.pack(fill="x")

        av_size = 52  # Tamaño del avatar grande en el menú
        av = tk.Canvas(top, width=av_size, height=av_size, bg=BG2, highlightthickness=0, cursor="hand2")
        av.pack(side="left")

        fp = self.usuario.get("foto_path")
        try:
            from PIL import Image, ImageTk, ImageDraw
            if not fp or not os.path.exists(fp):
                raise FileNotFoundError
            img = Image.open(fp).convert("RGBA").resize((av_size, av_size), Image.LANCZOS)
            mask = Image.new("L", (av_size, av_size), 0)
            ImageDraw.Draw(mask).ellipse((0, 0, av_size, av_size), fill=255)
            img.putalpha(mask)
            self._menu_foto_ref = ImageTk.PhotoImage(img)  # Referencia de imagen en el menú
            av.create_image(av_size // 2, av_size // 2, image=self._menu_foto_ref)
            av.create_oval(1, 1, av_size - 1, av_size - 1, outline=ACC, width=2)
        except Exception:
            self._menu_foto_ref = None
            av.create_oval(1, 1, av_size - 1, av_size - 1, fill="#4C1D95", outline=ACC, width=2)
            av.create_text(av_size // 2, av_size // 2,
                           text=self.usuario["nombre"][0].upper(),
                           fill="white", font=("Segoe UI", 20, "bold"))

        def cambiar_foto(e=None):
            """Cierra el menú y abre el diálogo de cambio de foto."""
            self._on_menu_close(menu)
            self._cambiar_foto_rapido()

        av.bind("<Button-1>", cambiar_foto)  # Clic en el avatar del menú cambia la foto

        # Información del usuario al lado del avatar
        info = tk.Frame(top, bg=BG2)
        info.pack(side="left", padx=(12, 0))
        tk.Label(info, text=self.usuario["nombre"], bg=BG2, fg="#EDE9FE", font=("Segoe UI", 11, "bold")).pack(anchor="w")
        tk.Label(info, text=self.usuario["correo"],  bg=BG2, fg="#6D6A9C", font=("Segoe UI", 8)).pack(anchor="w", pady=(3, 0))
        lbl_cf = tk.Label(info, text="✎ Cambiar foto", bg=BG2, fg=ACC, font=("Segoe UI", 8, "underline"), cursor="hand2")
        lbl_cf.pack(anchor="w", pady=(5, 0))
        lbl_cf.bind("<Button-1>", cambiar_foto)

        tk.Frame(outer, bg=SEP, height=1).pack(fill="x")  # Separador horizontal

        def item(icon, label, cmd):
            """Crea un ítem de menú estándar con ícono y texto."""
            f = tk.Frame(outer, bg=BG, cursor="hand2", padx=16, pady=10)
            f.pack(fill="x")
            tk.Label(f, text=icon,  bg=BG, fg="#A78BFA", font=("Segoe UI", 12)).pack(side="left", padx=(0, 12))
            tk.Label(f, text=label, bg=BG, fg="#D1D5DB", font=("Segoe UI", 10)).pack(side="left")
            def clk(e, _cmd=cmd):
                self._on_menu_close(menu)
                _cmd()
            for w in (f, *f.winfo_children()):
                w.bind("<Button-1>", clk)  # Clic en cualquier parte del ítem ejecuta el comando

        item("✏️", "Editar perfil", self._editar_perfil)
        tk.Frame(outer, bg=SEP, height=1).pack(fill="x")

        def item_danger(icon, label, cmd):
            """Crea un ítem de menú de acción peligrosa (en rojo)."""
            f = tk.Frame(outer, bg=BG, cursor="hand2", padx=16, pady=10)
            f.pack(fill="x")
            tk.Label(f, text=icon,  bg=BG, fg="#F87171", font=("Segoe UI", 12)).pack(side="left", padx=(0, 12))
            tk.Label(f, text=label, bg=BG, fg="#F87171", font=("Segoe UI", 10)).pack(side="left")
            def clk(e, _cmd=cmd):
                self._on_menu_close(menu)
                _cmd()
            for w in (f, *f.winfo_children()):
                w.bind("<Button-1>", clk)

        item_danger("⏻", "Cerrar sesión", self._logout)

        # Cierra el menú al perder el foco, presionar Escape o mover la ventana
        def _focus_out(e):
            self.root.after(100, lambda: self._on_menu_close(menu) if self._menu_abierto else None)

        menu.bind("<FocusOut>", _focus_out)
        menu.bind("<Escape>", lambda e: self._on_menu_close(menu))
        self.root.bind("<Configure>", lambda e: self._on_menu_close(menu), add="+")
        menu.focus_set()

    def _on_menu_close(self, menu):
        """
        Cierra el menú de perfil y restaura el estado visual del avatar.
        Desvincula el evento <Configure> de la ventana raíz para evitar múltiples cierres.
        """
        try:
            self.root.unbind("<Configure>")
        except Exception:
            pass

        self._menu_abierto = None

        try:
            if menu.winfo_exists():
                menu.destroy()
        except Exception:
            pass

        # Cancela animación de hover pendiente
        if self._hover_anim_id:
            try:
                self.root.after_cancel(self._hover_anim_id)
            except Exception:
                pass
            self._hover_anim_id = None

        # Restaura el color de fondo original del avatar
        self._pill_bg = BG_HEADER
        self._pill_canvas.config(bg=BG_HEADER, highlightthickness=0)
        self._dibujar_pill_canvas()

    def _cambiar_foto_rapido(self):
        """
        Permite cambiar la foto de perfil directamente desde el menú,
        sin abrir la ventana de edición completa.
        Actualiza la BD y redibuja el avatar.
        """
        path = filedialog.askopenfilename(
            title="Elegir foto",
            filetypes=[("Imágenes", "*.png *.jpg *.jpeg"), ("Todos", "*.*")]
        )
        if not path:
            return
        actualizar_usuario(self.usuario["id"], self.usuario["nombre"], self.usuario["correo"], path)
        self.usuario["foto_path"] = path  # Actualiza el diccionario en memoria
        self._dibujar_pill_canvas()       # Redibuja el avatar con la nueva foto

    def _logout(self):
        """
        Cierra la sesión del usuario tras confirmación.
        Limpia la ventana principal y regresa a la pantalla de login.
        """
        ok = Messagebox.yesno("¿Cerrar sesión?", "Salir", parent=self)
        if ok == "Yes":
            for w in self.root.winfo_children():
                w.destroy()
            self.root.withdraw()  # Oculta la ventana principal
            abrir_login()         # Abre la ventana de login de nuevo

    def _editar_perfil(self):
        """Abre la ventana de edición de perfil."""
        EditarPerfilWindow(self, self.usuario, self._on_perfil_actualizado)

    def _on_perfil_actualizado(self, u):
        """
        Callback que recibe los datos actualizados del perfil.
        Si u es None, significa que la cuenta fue eliminada → vuelve al login.
        Si u tiene datos, actualiza el diccionario local y redibuja el avatar.
        """
        if u is None:
            # Cuenta eliminada: limpia la UI y vuelve al login
            for w in self.root.winfo_children():
                w.destroy()
            self.root.withdraw()
            abrir_login()
            return
        self.usuario = u              # Actualiza los datos en memoria
        self._dibujar_pill_canvas()   # Redibuja el avatar con los nuevos datos

    # ── Menú ⋮ ────────────────────────────────────────────

    def _toggle_dot_menu(self, event=None):
        """Alterna la visibilidad del menú ⋮. Si está abierto lo cierra, si no lo abre."""
        if self._menu_dot_abierto:
            try:
                if self._menu_dot_abierto.winfo_exists():
                    self._cerrar_dot_menu()
                    return
            except Exception:
                pass
        self._show_dot_menu()

    def _show_dot_menu(self):
        """
        Muestra el menú ⋮ con opciones: Exportar Excel, Exportar PDF, Eliminar seleccionado.
        Se posiciona en la esquina superior derecha de la ventana, debajo del encabezado.
        """
        BG  = "#0D0C22"
        SEP = "#1E1C42"
        ACC = "#7C3AED"

        menu = tk.Toplevel(self.root)
        menu.overrideredirect(True)
        menu.attributes("-topmost", True)

        rw = 220  # Ancho del menú
        rx = self.root.winfo_rootx() + self.root.winfo_width() - rw - 8
        ry = self.root.winfo_rooty() + 62  # Posicionado debajo del encabezado
        menu.geometry(f"+{rx}+{ry}")

        self._menu_dot_abierto = menu

        outer = tk.Frame(menu, bg=BG, highlightbackground=ACC, highlightthickness=1)
        outer.pack(fill="both", expand=True)

        # Ítems del menú: (ícono, etiqueta, comando)
        items = [
            ("📊", "Exportar a Excel",       self._exportar_excel),
            ("📄", "Exportar a PDF",          self._exportar_pdf),
            (None, None, None),               # Separador visual
            ("🗑️", "Eliminar seleccionado",  self._eliminar_seleccionado),
        ]

        for icon, label, cmd in items:
            if icon is None:
                tk.Frame(outer, bg=SEP, height=1).pack(fill="x")  # Línea separadora
                continue

            # Color diferente para la opción de eliminar (rojo de peligro)
            fg_i = "#A78BFA" if "Elim" not in label else "#F87171"
            fg_l = "#D1D5DB" if "Elim" not in label else "#F87171"

            f = tk.Frame(outer, bg=BG, cursor="hand2", padx=14, pady=9)
            f.pack(fill="x")
            tk.Label(f, text=icon,  bg=BG, fg=fg_i, font=("Segoe UI", 11)).pack(side="left", padx=(0, 10))
            tk.Label(f, text=label, bg=BG, fg=fg_l, font=("Segoe UI", 10)).pack(side="left")

            def clk(e, _cmd=cmd):
                self._cerrar_dot_menu()
                _cmd()
            for w in (f, *f.winfo_children()):
                w.bind("<Button-1>", clk)

        # Cierra al perder foco, Escape o mover la ventana
        def _focus_out(e):
            self.root.after(100, lambda: self._cerrar_dot_menu() if self._menu_dot_abierto else None)

        menu.bind("<FocusOut>", _focus_out)
        menu.bind("<Escape>", lambda e: self._cerrar_dot_menu())
        self.root.bind("<Configure>", lambda e: self._cerrar_dot_menu(), add="+")
        menu.focus_set()

    def _cerrar_dot_menu(self):
        """Cierra el menú ⋮ y limpia la referencia."""
        try:
            self.root.unbind("<Configure>")
        except Exception:
            pass
        if self._menu_dot_abierto:
            try:
                if self._menu_dot_abierto.winfo_exists():
                    self._menu_dot_abierto.destroy()
            except Exception:
                pass
            self._menu_dot_abierto = None

    # ── Actualización de datos ────────────────────────────

    def _auto_refresh(self):
        """
        Programa una actualización automática de los datos cada 60 segundos.
        Usa after() para programar la siguiente llamada sin bloquear la UI.
        """
        self._refresh()
        self.root.after(60_000, self._auto_refresh)  # 60,000 ms = 60 segundos

    def _refresh(self):
        """
        Actualiza todos los componentes visuales con los datos más recientes de la BD:
        combo de meses, tabla de movimientos, tarjetas de resumen y gráficas.
        """
        self._actualizar_meses()
        self._cargar_tabla()
        self._actualizar_resumen()
        self._animar_graficas()

    def _actualizar_meses(self):
        """
        Actualiza las opciones del ComboBox de meses con los meses disponibles en la BD.
        Si el mes seleccionado ya no existe, selecciona el más reciente.
        """
        meses = meses_disponibles(self.usuario["id"])
        self.cb_mes["values"] = meses

        if meses and self.mes_var.get() not in meses:
            self.mes_var.set(meses[0])   # Selecciona el mes más reciente por defecto
        elif not meses:
            self.mes_var.set("")          # Limpia si no hay meses

    def _rango_actual(self):
        """
        Devuelve el rango de fechas según el mes seleccionado en el ComboBox.
        Si no hay mes seleccionado (Ver todo), devuelve un rango que abarca todo.
        """
        mes = self.mes_var.get()
        return rango_del_mes(mes) if mes else ("0000-00-00 00:00:00", "9999-99-99 23:59:59")

    def _ver_todo(self):
        """Limpia el filtro de mes para mostrar todos los movimientos y refresca."""
        self.mes_var.set("")
        self._refresh()

    def _cargar_tabla(self):
        """
        Recarga el Treeview con los movimientos del rango de fechas actual.
        Limpia las filas existentes antes de insertar las nuevas.
        Aplica tags de color según tipo (Ingreso=verde, Gasto=rojo).
        """
        for row in self.tree.get_children():
            self.tree.delete(row)  # Limpia la tabla

        inicio, fin = self._rango_actual()
        for mid, tipo, concepto, monto, fh in obtener_movimientos_rango(self.usuario["id"], inicio, fin):
            self.tree.insert("", "end",
                             values=(mid, tipo, concepto, f"${monto:,.2f}", fh),
                             tags=(tipo,))  # El tag determina el color de la fila

    def _actualizar_resumen(self):
        """
        Actualiza el texto y color de las tarjetas de resumen (Ingresos, Gastos, Saldo).
        Si el saldo es negativo, lo muestra en rojo en lugar de azul.
        """
        inicio, fin = self._rango_actual()
        ing, gas, saldo = calcular_totales_rango(self.usuario["id"], inicio, fin)

        self.lbl_ing.config(text=f"Ingresos: ${ing:,.2f}",   fg=GREEN)
        self.lbl_gas.config(text=f"Gastos: ${gas:,.2f}",     fg=RED)
        self.lbl_saldo.config(text=f"Saldo: ${saldo:,.2f}",  fg=BLUE if saldo >= 0 else RED)

    # ── Gráficas ──────────────────────────────────────────

    def _animar_graficas(self):
        """
        Inicia la animación de las barras del resumen del período.
        Resetea los valores actuales a 0 y configura los valores objetivo (target).
        También dibuja las líneas históricas (sin animación).
        """
        inicio, fin = self._rango_actual()
        ing, gas, saldo = calcular_totales_rango(self.usuario["id"], inicio, fin)

        self._anim_target  = [ing, gas, saldo]
        self._anim_current = [0.0, 0.0, 0.0]
        self._anim_step    = 0

        self._dibujar_lineas_historicas()  # Dibuja el gráfico de líneas (estático)

        # Cancela animación previa si estaba en curso
        if self._anim_id:
            try:
                self.root.after_cancel(self._anim_id)
            except Exception:
                pass

        self._tick_bar_anim()  # Inicia el primer tick de animación

    def _tick_bar_anim(self):
        """
        Ejecuta un paso de la animación de barras usando easing cúbico (ease-out).
        Se llama a sí misma cada 30ms hasta completar STEPS pasos.
        """
        STEPS = 20
        self._anim_step += 1

        ease = 1 - (1 - self._anim_step / STEPS) ** 3  # Easing cúbico ease-out
        self._anim_current = [v * ease for v in self._anim_target]

        self._dibujar_barras(*self._anim_current)
        self.canvas_fig.draw_idle()  # Redibuja el canvas de matplotlib de forma eficiente

        if self._anim_step < STEPS:
            self._anim_id = self.root.after(30, self._tick_bar_anim)  # Siguiente frame en 30ms

    def _dibujar_barras(self, ing, gas, saldo):
        """
        Dibuja la gráfica de barras del período actual con efecto de brillo (glow).
        Muestra tres barras: Ingresos (verde), Gastos (rojo), Saldo (azul o rojo si negativo).
        Incluye etiquetas de valor encima de cada barra y líneas de brillo.
        """
        BG   = "#0F172A"  # Fondo oscuro
        GRID = "#1E293B"  # Color de las líneas de grilla

        ax = self.ax_bar
        ax.clear()
        ax.set_facecolor(BG)

        etiquetas = ["Ingresos", "Gastos", "Saldo"]
        valores   = [ing, gas, saldo]
        bcolors   = ["#10B981", "#EF4444", "#3B82F6" if saldo >= 0 else "#EF4444"]
        glow      = ["#34D399", "#F87171", "#60A5FA" if saldo >= 0 else "#F87171"]  # Colores de brillo

        x = np.arange(len(etiquetas))

        # Barras fantasma con baja opacidad para simular el efecto glow
        for xi, val, gc in zip(x, valores, glow):
            ax.bar(xi, val, width=0.62, color=gc, alpha=0.18, zorder=1)

        # Barras principales
        bars = ax.bar(x, valores, width=0.48, color=bcolors, edgecolor="none", zorder=3)

        # Línea horizontal en el tope de cada barra para refuerzo visual
        for bar, gc in zip(bars, glow):
            h  = bar.get_height()
            xi = bar.get_x()
            w  = bar.get_width()
            ax.plot([xi, xi + w], [h, h], color=gc, linewidth=2, zorder=4)

        mx = max(abs(v) for v in valores) if any(valores) else 1

        # Etiquetas de valor encima de cada barra
        for bar, val, gc in zip(bars, valores, glow):
            ax.text(bar.get_x() + bar.get_width() / 2,
                    bar.get_height() + mx * 0.04,
                    f"${val:,.0f}",
                    ha="center", va="bottom",
                    fontsize=8, fontweight="bold", color=gc, zorder=5)

        # Configuración visual de ejes y grilla
        ax.set_xticks(x)
        ax.set_xticklabels(etiquetas, fontsize=9, color="#94A3B8", fontweight="bold")
        ax.tick_params(axis="y", labelsize=7.5, colors="#475569")
        ax.set_title("Resumen del período", fontsize=10, fontweight="bold", color="#E2E8F0", pad=8)

        for spine in ax.spines.values():
            spine.set_color(GRID)

        ax.set_axisbelow(True)
        ax.yaxis.grid(True, color=GRID, linewidth=0.8, linestyle="--")
        ax.xaxis.grid(False)

        if mx > 0:
            ax.set_ylim(bottom=min(0, min(valores)) - mx * 0.05,
                        top=max(valores) + mx * 0.22)

    def _dibujar_lineas_historicas(self):
        """
        Dibuja la gráfica de líneas con el historial mensual del usuario.
        Muestra tres series: Ingresos, Gastos y Saldo con marcadores y áreas rellenas.
        Si no hay datos históricos, muestra un mensaje informativo.
        """
        BG   = "#0F172A"
        GRID = "#1E293B"

        ax = self.ax_line
        ax.clear()
        ax.set_facecolor(BG)

        meses = sorted(meses_disponibles(self.usuario["id"]))  # Meses ordenados cronológicamente

        if not meses:
            # Mensaje cuando no hay datos históricos
            ax.text(0.5, 0.5,
                    "Sin historial aún\nRegistra movimientos para ver el historial",
                    ha="center", va="center", transform=ax.transAxes,
                    fontsize=9, color="#475569", multialignment="center")
            ax.set_title("Historial mensual", fontsize=10, fontweight="bold", color="#94A3B8")
            for spine in ax.spines.values():
                spine.set_color(GRID)
            return

        # Convierte "YYYY-MM" a etiquetas legibles como "Ene 25"
        etiquetas_x = []
        for m in meses:
            try:
                etiquetas_x.append(datetime.strptime(m, "%Y-%m").strftime("%b %y"))
            except Exception:
                etiquetas_x.append(m[5:] + "/" + m[2:4])

        # Calcula totales por cada mes histórico
        ing_vals = []
        gas_vals = []
        sal_vals = []
        for m in meses:
            ini, fin = rango_del_mes(m)
            i, g, s  = calcular_totales_rango(self.usuario["id"], ini, fin)
            ing_vals.append(i)
            gas_vals.append(g)
            sal_vals.append(s)

        x = np.arange(len(meses))

        # Serie de Ingresos: área rellena + línea con puntos
        ax.fill_between(x, ing_vals, alpha=0.15, color="#10B981", interpolate=True)
        ax.plot(x, ing_vals, color="#10B981", linewidth=2.5, marker="o", markersize=7,
                markerfacecolor="#0F172A", markeredgecolor="#10B981", markeredgewidth=2.2,
                label="Ingresos", zorder=5)

        # Serie de Gastos: área rellena + línea con puntos
        ax.fill_between(x, gas_vals, alpha=0.12, color="#EF4444", interpolate=True)
        ax.plot(x, gas_vals, color="#EF4444", linewidth=2.2, marker="o", markersize=7,
                markerfacecolor="#0F172A", markeredgecolor="#EF4444", markeredgewidth=2.2,
                label="Gastos", zorder=5)

        # Serie de Saldo: línea punteada con triángulos
        ax.plot(x, sal_vals, color="#3B82F6", linewidth=2.0, linestyle="--", marker="^", markersize=7,
                markerfacecolor="#0F172A", markeredgecolor="#3B82F6", markeredgewidth=2,
                label="Saldo", zorder=5)

        # Etiquetas de valor encima de cada punto
        for xi, (iv, gv, sv) in enumerate(zip(ing_vals, gas_vals, sal_vals)):
            for val, col in [(iv, "#34D399"), (gv, "#F87171"), (sv, "#60A5FA")]:
                lbl = f"${val/1000:.0f}k" if abs(val) >= 1000 else f"${val:.0f}"
                ax.annotate(lbl, xy=(xi, val), xytext=(0, 9), textcoords="offset points",
                            ha="center", fontsize=7.5, fontweight="bold", color=col)

        # Línea cero si hay saldos negativos
        if min(sal_vals) < 0:
            ax.axhline(0, color="#334155", linewidth=1, linestyle=":")

        # Configuración de ejes y leyenda
        ax.set_xticks(x)
        ax.set_xticklabels(etiquetas_x, fontsize=8.5, color="#94A3B8", fontweight="bold")
        ax.tick_params(axis="y", labelsize=7.5, colors="#475569")
        ax.set_title("Historial mensual", fontsize=10, fontweight="bold", color="#E2E8F0", pad=8)

        for spine in ax.spines.values():
            spine.set_color(GRID)

        ax.set_axisbelow(True)
        ax.yaxis.grid(True, color=GRID, linewidth=0.7, linestyle="--")
        ax.xaxis.grid(True, color=GRID, linewidth=0.3, linestyle=":")

        # Ajusta el límite del eje Y con padding proporcional
        all_vals = ing_vals + gas_vals + sal_vals
        if all_vals:
            top = max(all_vals)
            bot = min(all_vals)
            pad = (top - bot) * 0.22 if top != bot else abs(top) * 0.3 or 1
            ax.set_ylim(bot - pad * 0.3, top + pad)

        # Leyenda con estilo oscuro
        leg = ax.legend(loc="upper left", fontsize=8, framealpha=0.25,
                        edgecolor="#334155", facecolor="#1E293B",
                        handlelength=1.8, markerscale=0.9)
        for txt in leg.get_texts():
            txt.set_color("#E2E8F0")

    # ── Registrar y eliminar movimientos ──────────────────

    def _registrar(self):
        """
        Valida y registra un nuevo movimiento (ingreso o gasto).
        Para gastos, verifica que el monto no supere el saldo disponible.
        Limpia los campos tras el registro exitoso y refresca la vista.
        """
        tipo      = self.tipo_var.get()
        concepto  = self.concepto_var.get().strip()
        monto_s   = self.monto_var.get().strip().replace(",", ".")  # Acepta coma decimal

        if not concepto:
            Messagebox.show_warning("Escribe un concepto.", "Atención", parent=self)
            return

        try:
            monto = float(monto_s)
            if monto <= 0:
                raise ValueError
        except ValueError:
            Messagebox.show_error("Monto inválido.", "Error", parent=self)
            return

        # Validación de saldo suficiente solo para gastos
        if tipo == "Gasto":
            inicio, fin = self._rango_actual()
            ing, gas, saldo = calcular_totales_rango(self.usuario["id"], inicio, fin)
            if monto > saldo:
                Messagebox.show_error(
                    f"Saldo insuficiente.\nDisponible: ${saldo:,.2f}\nIntentaste: ${monto:,.2f}",
                    "Sin fondos", parent=self
                )
                return

        insertar_movimiento(self.usuario["id"], tipo, concepto, monto)
        self.concepto_var.set("")  # Limpia los campos del formulario
        self.monto_var.set("")
        self._refresh()

    def _eliminar_seleccionado(self):
        """
        Elimina el movimiento seleccionado en el Treeview tras confirmación del usuario.
        Si no hay nada seleccionado, muestra una advertencia.
        """
        sel = self.tree.selection()
        if not sel:
            Messagebox.show_warning("Selecciona un movimiento.", "Atención", parent=self)
            return

        mov_id = self.tree.item(sel[0])["values"][0]  # Obtiene el ID del movimiento seleccionado
        ok = Messagebox.yesno(f"¿Eliminar movimiento #{mov_id}?", "Confirmar")
        if ok == "Yes":
            eliminar_movimiento_por_id(mov_id, self.usuario["id"])
            self._refresh()

    # ── Exportación ───────────────────────────────────────

    def _generar_imagen_grafica(self):
        """
        Exporta las gráficas actuales como imagen JPEG temporal.
        Usa PIL si está disponible para mayor calidad. Si no, usa matplotlib directamente.
        Devuelve la ruta del archivo temporal creado.
        """
        tmp_fd, tmp_path = tempfile.mkstemp(suffix=".jpg")
        os.close(tmp_fd)  # Cierra el file descriptor antes de escritura
        try:
            from PIL import Image as PILImage
            import io
            buf = io.BytesIO()
            self.fig.savefig(buf, dpi=130, bbox_inches="tight",
                             facecolor=self.fig.get_facecolor(), format="png")
            buf.seek(0)
            PILImage.open(buf).convert("RGB").save(tmp_path, format="JPEG", quality=95)
        except ImportError:
            # Fallback sin PIL
            self.fig.savefig(tmp_path, dpi=130, bbox_inches="tight", facecolor="#0F172A", format="jpeg")
        return tmp_path

    def _nombre_extracto(self):
        """
        Genera el nombre base del archivo de exportación.
        Formato: Extracto_NombreUsuario_YYYY-MM
        """
        mes = self.mes_var.get() or datetime.now().strftime("%Y-%m")
        return f"Extracto_{self.usuario['nombre'].replace(' ', '_')}_{mes}"

    def _exportar_excel(self):
        """
        Exporta los movimientos del período actual a un archivo Excel (.xlsx).
        Incluye:
        - Hoja "Movimientos": encabezado, resumen de totales y tabla de movimientos con colores.
        - Hoja "Gráficas": imagen de las gráficas actuales.
        Muestra un diálogo de guardado para que el usuario elija la ubicación.
        """
        inicio, fin = self._rango_actual()
        filas = obtener_movimientos_rango(self.usuario["id"], inicio, fin)
        ing, gas, saldo = calcular_totales_rango(self.usuario["id"], inicio, fin)

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=self._nombre_extracto() + ".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            title="Guardar extracto Excel"
        )
        if not path:
            return  # Usuario canceló el diálogo

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Movimientos"

            # Fila 1: Título del extracto
            ws.merge_cells("A1:E1")
            ws["A1"] = f"Extracto — {self.usuario['nombre']}"
            ws["A1"].font      = Font(bold=True, size=14, color="7C3AED")
            ws["A1"].alignment = Alignment(horizontal="center")

            # Fila 2: Período y fecha de generación
            ws.merge_cells("A2:E2")
            ws["A2"] = f"Período: {inicio[:10]} → {fin[:10]}  |  {datetime.now().strftime('%d/%m/%Y %H:%M')}"
            ws["A2"].font      = Font(italic=True, size=10, color="6B7280")
            ws["A2"].alignment = Alignment(horizontal="center")

            ws.merge_cells("A3:E3")  # Fila vacía separadora
            ws["A3"] = ""

            # Fila 4: Encabezado de la sección Resumen
            ws.merge_cells("A4:B4")
            ws["A4"] = "Resumen"
            ws["A4"].font      = Font(bold=True, size=11, color="FFFFFF")
            ws["A4"].fill      = PatternFill("solid", fgColor="7C3AED")
            ws["A4"].alignment = Alignment(horizontal="center")

            # Filas 5-7: Totales de ingresos, gastos y saldo con color por tipo
            for off, (lbl, val, color) in enumerate([
                ("Ingresos", ing,   "059669"),
                ("Gastos",   gas,   "DC2626"),
                ("Saldo",    saldo, "2563EB")
            ]):
                r = 5 + off
                ws.cell(row=r, column=1, value=lbl).font  = Font(bold=True, color=color)
                c = ws.cell(row=r, column=2, value=val)
                c.number_format = '"$"#,##0.00'
                c.font = Font(bold=True, color=color)

            # Fila 10: Encabezados de columnas de la tabla
            fill_h = PatternFill("solid", fgColor="4C1D95")
            for col, h in enumerate(["ID", "Tipo", "Concepto", "Monto", "Fecha / Hora"], 1):
                cell = ws.cell(row=10, column=col, value=h)
                cell.font      = Font(bold=True, color="FFFFFF")
                cell.fill      = fill_h
                cell.alignment = Alignment(horizontal="center")

            # Filas de datos con color alternado según tipo (verde=ingreso, rojo=gasto)
            fill_ing = PatternFill("solid", fgColor="D1FAE5")
            fill_gas = PatternFill("solid", fgColor="FEE2E2")
            for r, (mid, tipo, concepto, monto, fh) in enumerate(filas, 11):
                ws.cell(row=r, column=1, value=mid)
                ws.cell(row=r, column=2, value=tipo)
                ws.cell(row=r, column=3, value=concepto)
                ws.cell(row=r, column=4, value=monto).number_format = '"$"#,##0.00'
                ws.cell(row=r, column=5, value=fh)
                f = fill_ing if tipo == "Ingreso" else fill_gas
                for c in range(1, 6):
                    ws.cell(row=r, column=c).fill = f

            # Ancho de columnas ajustado al contenido esperado
            for col_w, width in zip("ABCDE", [8, 12, 38, 16, 22]):
                ws.column_dimensions[col_w].width = width

            # Hoja de gráficas
            ws2 = wb.create_sheet("Gráficas")
            ws2.merge_cells("A1:J1")
            ws2["A1"] = f"Gráficas — {self.usuario['nombre']}"
            ws2["A1"].font      = Font(bold=True, size=13, color="7C3AED")
            ws2["A1"].alignment = Alignment(horizontal="center")

            img_path = self._generar_imagen_grafica()

            try:
                xl_img = XLImage(img_path)
                xl_img.anchor = "A3"
                xl_img.width  = 700
                xl_img.height = 500
                ws2.add_image(xl_img)
                wb.save(path)
                Messagebox.show_info(f"✅ Excel guardado:\n{path}", "Exportado")
            except Exception as e:
                Messagebox.show_error(f"Error Excel:\n{e}", "Error")
            finally:
                try:
                    os.remove(img_path)  # Elimina la imagen temporal
                except Exception:
                    pass

        except PermissionError:
            # El archivo está abierto en Excel y no se puede sobreescribir
            Messagebox.show_error("Archivo abierto en Excel. Ciérralo e intenta de nuevo.", "Permiso denegado")
        except Exception as e:
            Messagebox.show_error(f"Error:\n{e}", "Error")

    def _exportar_pdf(self):
        """
        Exporta los movimientos del período actual a un archivo PDF con ReportLab.
        Incluye:
        - Encabezado con nombre del usuario y período.
        - Tabla de resumen (Ingresos, Gastos, Saldo).
        - Tabla detallada de movimientos con colores por tipo.
        - Segunda página con las gráficas actuales como imagen.
        """
        inicio, fin = self._rango_actual()
        filas = obtener_movimientos_rango(self.usuario["id"], inicio, fin)
        ing, gas, saldo = calcular_totales_rango(self.usuario["id"], inicio, fin)

        path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            initialfile=self._nombre_extracto() + ".pdf",
            filetypes=[("PDF", "*.pdf")],
            title="Guardar extracto PDF"
        )
        if not path:
            return

        try:
            # Documento PDF con márgenes de 1.8cm en los lados
            doc = SimpleDocTemplate(path, pagesize=A4,
                                    leftMargin=1.8*cm, rightMargin=1.8*cm,
                                    topMargin=2*cm, bottomMargin=2*cm)
            styles = getSampleStyleSheet()
            story  = []  # Lista de elementos del documento

            # Estilos personalizados
            title_s = ParagraphStyle("t", parent=styles["Title"],
                                     textColor=colors.HexColor("#7C3AED"), fontSize=18, spaceAfter=4)
            sub_s   = ParagraphStyle("s", parent=styles["Normal"],
                                     textColor=colors.HexColor("#6B7280"), fontSize=10, spaceAfter=2, alignment=1)

            # Encabezado
            story.append(Paragraph("Extracto de Presupuesto", title_s))
            story.append(Paragraph(self.usuario["nombre"], sub_s))
            story.append(Paragraph(
                f"Período: <b>{inicio[:10]}</b> → <b>{fin[:10]}</b>  |  {datetime.now().strftime('%d/%m/%Y %H:%M')}",
                sub_s
            ))
            story.append(Spacer(1, 0.5*cm))

            # Tabla de resumen con colores por fila
            t_res = PDFTable(
                [["Concepto", "Monto"],
                 ["Ingresos totales",   f"${ing:,.2f}"],
                 ["Gastos totales",     f"${gas:,.2f}"],
                 ["Saldo disponible",   f"${saldo:,.2f}"]],
                colWidths=[8*cm, 5*cm]
            )
            t_res.setStyle(TableStyle([
                ("BACKGROUND",   (0, 0), (-1, 0), colors.HexColor("#4C1D95")),
                ("TEXTCOLOR",    (0, 0), (-1, 0), colors.white),
                ("FONTNAME",     (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE",     (0, 0), (-1, -1), 10),
                ("ALIGN",        (1, 0), (1, -1), "RIGHT"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F3FF")]),
                ("TEXTCOLOR",    (0, 1), (0, 1), colors.HexColor("#059669")),  # Verde ingresos
                ("TEXTCOLOR",    (0, 2), (0, 2), colors.HexColor("#DC2626")),  # Rojo gastos
                ("TEXTCOLOR",    (0, 3), (0, 3), colors.HexColor("#2563EB")),  # Azul saldo
                ("FONTNAME",     (0, 1), (-1, -1), "Helvetica-Bold"),
                ("BOX",          (0, 0), (-1, -1), 0.8, colors.HexColor("#7C3AED")),
                ("INNERGRID",    (0, 0), (-1, -1), 0.25, colors.HexColor("#E5E7EB")),
                ("TOPPADDING",   (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING",(0, 0), (-1, -1), 6),
                ("LEFTPADDING",  (0, 0), (-1, -1), 10),
            ]))
            story.append(t_res)
            story.append(Spacer(1, 0.6*cm))

            # Tabla de detalle de movimientos
            if filas:
                story.append(Paragraph("Detalle de movimientos",
                    ParagraphStyle("h3", parent=styles["Heading3"],
                                   textColor=colors.HexColor("#4C1D95"))))
                story.append(Spacer(1, 0.2*cm))

                data = [["ID", "Tipo", "Concepto", "Monto", "Fecha"]] + [
                    [str(mid), tipo, concepto, f"${monto:,.2f}", fh[:16]]
                    for mid, tipo, concepto, monto, fh in filas
                ]
                t = PDFTable(data, colWidths=[1.2*cm, 2.5*cm, 7*cm, 3*cm, 4*cm])
                ts2 = TableStyle([
                    ("BACKGROUND",   (0, 0), (-1, 0), colors.HexColor("#4C1D95")),
                    ("TEXTCOLOR",    (0, 0), (-1, 0), colors.white),
                    ("FONTNAME",     (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE",     (0, 0), (-1, -1), 8),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
                    ("BOX",          (0, 0), (-1, -1), 0.5, colors.HexColor("#D1D5DB")),
                    ("INNERGRID",    (0, 0), (-1, -1), 0.25, colors.HexColor("#E5E7EB")),
                    ("ALIGN",        (3, 0), (3, -1), "RIGHT"),
                    ("TOPPADDING",   (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING",(0, 0), (-1, -1), 4),
                ])

                # Colorea filas individuales según tipo (verde=ingreso, rojo=gasto)
                for i, (_, tipo, *_r) in enumerate(filas, 1):
                    ts2.add("BACKGROUND", (0, i), (-1, i),
                            colors.HexColor("#D1FAE5") if tipo == "Ingreso" else colors.HexColor("#FEE2E2"))

                t.setStyle(ts2)
                story.append(t)
            else:
                story.append(Paragraph("No hay movimientos en este período.", styles["Normal"]))

            # Segunda página: gráficas
            story.append(PageBreak())
            story.append(Paragraph("Gráficas del período",
                ParagraphStyle("gt", parent=styles["Heading2"],
                               textColor=colors.HexColor("#7C3AED"))))
            story.append(Spacer(1, 0.4*cm))

            img_path = self._generar_imagen_grafica()
            try:
                page_w = A4[0] - 3.6*cm  # Ancho disponible en la página
                story.append(PDFImage(img_path, width=page_w, height=page_w * 0.72))
                doc.build(story)
                Messagebox.show_info(f"✅ PDF guardado:\n{path}", "Exportado")
            except Exception as e:
                Messagebox.show_error(f"Error PDF:\n{e}", "Error")
            finally:
                try:
                    os.remove(img_path)  # Limpia la imagen temporal
                except Exception:
                    pass

        except Exception as e:
            Messagebox.show_error(f"Error:\n{e}", "Error")


# ============================================================
# MAIN — Punto de entrada de la aplicación
# ============================================================

def iniciar_app(usuario):
    """
    Limpia todos los widgets de la ventana raíz y crea la instancia principal de App
    con los datos del usuario autenticado.
    """
    for w in root.winfo_children():
        w.destroy()
    App(root, usuario)


def abrir_login():
    """
    Abre la ventana de inicio de sesión.
    Configura el protocolo de cierre para destruir la ventana raíz si se cierra el login.
    """
    win = LoginWindow(root, _on_login_ok)
    win.protocol("WM_DELETE_WINDOW", root.destroy)  # Cerrar login = cerrar la app


def _on_login_ok(user):
    """
    Callback llamado por LoginWindow cuando el login fue exitoso.
    Muestra la ventana raíz (que estaba oculta) e inicia la app principal.
    """
    root.deiconify()  # Muestra la ventana raíz que estaba oculta con withdraw()
    iniciar_app(user)


# ── Inicialización ────────────────────────────────────────────
if __name__ == "__main__":
    crear_tabla_movimientos()   # Asegura que la tabla de movimientos existe en la BD
    crear_tabla_usuarios()      # Asegura que la tabla de usuarios existe en la BD

    root = tk.Tk()
    root.withdraw()  # Oculta la ventana raíz mientras se muestra el login

    # Configura el tema visual de ttk
    style = ttk.Style(root)
    try:
        style.theme_use("clam")  # Tema multiplataforma con mejor aspecto que el predeterminado
    except Exception:
        pass  # Si "clam" no está disponible, usa el tema por defecto

    # Estilos globales de fuente y padding para todos los widgets ttk
    style.configure(".",              font=("Segoe UI", 10))
    style.configure("TButton",        padding=6)
    style.configure("TEntry",         padding=4)
    style.configure("TLabelframe.Label", font=("Segoe UI", 10, "bold"))

    abrir_login()      # Inicia el flujo de autenticación
    root.mainloop()    # Inicia el loop de eventos de Tkinter (bloquea hasta cerrar la app)